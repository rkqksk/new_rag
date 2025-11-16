"""
Excel Export Service
Generate Excel reports for products, work orders, quality inspection
Version: v8.4.0
"""

import logging
from typing import List, Dict, Any, Optional
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart

logger = logging.getLogger(__name__)


class ExcelExportService:
    """Excel report generation service"""

    # Color palette
    COLORS = {
        'header': '4472C4',
        'success': '70AD47',
        'warning': 'FFC000',
        'danger': 'C00000',
        'info': '5B9BD5',
    }

    def __init__(self):
        """Initialize Excel export service"""
        self.workbook = None
        self.sheet = None

    def create_workbook(self, title: str = 'Report') -> Workbook:
        """Create new workbook"""
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = title
        return self.workbook

    def _apply_header_style(self, row: int, columns: int):
        """Apply header styling"""
        for col in range(1, columns + 1):
            cell = self.sheet.cell(row=row, column=col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['header'], end_color=self.COLORS['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'),
            )

    def _auto_adjust_column_width(self):
        """Auto-adjust column widths"""
        for column in self.sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max_length + 2, 50)
            self.sheet.column_dimensions[column_letter].width = adjusted_width

    async def export_products(
        self,
        products: List[Dict[str, Any]],
        include_images: bool = False
    ) -> bytes:
        """
        Export products to Excel

        Args:
            products: List of product dictionaries
            include_images: Include product images

        Returns:
            Excel file bytes
        """
        try:
            self.create_workbook('Products')

            # Headers
            headers = ['ID', 'Product Name', 'Code', 'Category', 'Material', 'Capacity', 'Price', 'Stock', 'MOQ', 'Supplier']
            for col, header in enumerate(headers, 1):
                self.sheet.cell(row=1, column=col, value=header)

            self._apply_header_style(1, len(headers))

            # Data rows
            for row, product in enumerate(products, 2):
                self.sheet.cell(row=row, column=1, value=product.get('id'))
                self.sheet.cell(row=row, column=2, value=product.get('product_name'))
                self.sheet.cell(row=row, column=3, value=product.get('product_code'))
                self.sheet.cell(row=row, column=4, value=product.get('category'))
                self.sheet.cell(row=row, column=5, value=product.get('material'))
                self.sheet.cell(row=row, column=6, value=product.get('capacity'))
                self.sheet.cell(row=row, column=7, value=product.get('price', 0))
                self.sheet.cell(row=row, column=8, value=product.get('stock', 0))
                self.sheet.cell(row=row, column=9, value=product.get('moq', 0))
                self.sheet.cell(row=row, column=10, value=product.get('supplier'))

            # Auto-adjust
            self._auto_adjust_column_width()

            # Summary
            summary_row = len(products) + 3
            self.sheet.cell(row=summary_row, column=1, value='Total Products:')
            self.sheet.cell(row=summary_row, column=2, value=len(products))
            self.sheet.cell(row=summary_row, column=1).font = Font(bold=True)

            # Save to bytes
            output = io.BytesIO()
            self.workbook.save(output)
            output.seek(0)

            logger.info(f'Exported {len(products)} products to Excel')
            return output.getvalue()

        except Exception as e:
            logger.error(f'Product export failed: {e}')
            raise

    async def export_work_orders(
        self,
        work_orders: List[Dict[str, Any]],
        include_charts: bool = True
    ) -> bytes:
        """
        Export work orders to Excel with charts

        Args:
            work_orders: List of work order dictionaries
            include_charts: Include progress charts

        Returns:
            Excel file bytes
        """
        try:
            self.create_workbook('Work Orders')

            # Headers
            headers = ['WO Number', 'Product', 'Quantity Planned', 'Quantity Completed', 'Status', 'Priority', 'Start Date', 'Due Date', 'Progress %']
            for col, header in enumerate(headers, 1):
                self.sheet.cell(row=1, column=col, value=header)

            self._apply_header_style(1, len(headers))

            # Data rows
            for row, wo in enumerate(work_orders, 2):
                self.sheet.cell(row=row, column=1, value=wo.get('wo_number'))
                self.sheet.cell(row=row, column=2, value=wo.get('product_name'))
                self.sheet.cell(row=row, column=3, value=wo.get('quantity_planned'))
                self.sheet.cell(row=row, column=4, value=wo.get('quantity_completed'))

                # Status with color
                status_cell = self.sheet.cell(row=row, column=5, value=wo.get('status'))
                status = wo.get('status', '').lower()
                if status == 'completed':
                    status_cell.fill = PatternFill(start_color=self.COLORS['success'], fill_type='solid')
                elif status == 'in_progress':
                    status_cell.fill = PatternFill(start_color=self.COLORS['info'], fill_type='solid')
                elif status == 'delayed':
                    status_cell.fill = PatternFill(start_color=self.COLORS['danger'], fill_type='solid')

                self.sheet.cell(row=row, column=6, value=wo.get('priority'))
                self.sheet.cell(row=row, column=7, value=wo.get('start_date'))
                self.sheet.cell(row=row, column=8, value=wo.get('due_date'))

                # Progress percentage
                progress = wo.get('progress_percent', 0)
                progress_cell = self.sheet.cell(row=row, column=9, value=progress)
                progress_cell.number_format = '0.00%'

                # Progress color
                if progress >= 100:
                    progress_cell.fill = PatternFill(start_color=self.COLORS['success'], fill_type='solid')
                elif progress >= 75:
                    progress_cell.fill = PatternFill(start_color=self.COLORS['info'], fill_type='solid')
                elif progress >= 50:
                    progress_cell.fill = PatternFill(start_color=self.COLORS['warning'], fill_type='solid')
                else:
                    progress_cell.fill = PatternFill(start_color=self.COLORS['danger'], fill_type='solid')

            # Auto-adjust
            self._auto_adjust_column_width()

            # Add chart if requested
            if include_charts and len(work_orders) > 0:
                chart_sheet = self.workbook.create_sheet('Progress Chart')
                chart = BarChart()
                chart.title = 'Work Order Progress'
                chart.x_axis.title = 'Work Orders'
                chart.y_axis.title = 'Progress %'

                data = Reference(self.sheet, min_col=9, min_row=1, max_row=len(work_orders) + 1)
                cats = Reference(self.sheet, min_col=1, min_row=2, max_row=len(work_orders) + 1)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)

                chart_sheet.add_chart(chart, 'A1')

            # Save to bytes
            output = io.BytesIO()
            self.workbook.save(output)
            output.seek(0)

            logger.info(f'Exported {len(work_orders)} work orders to Excel')
            return output.getvalue()

        except Exception as e:
            logger.error(f'Work order export failed: {e}')
            raise

    async def export_quality_report(
        self,
        inspections: List[Dict[str, Any]],
        summary_stats: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """
        Export quality inspection report

        Args:
            inspections: List of inspection results
            summary_stats: Summary statistics

        Returns:
            Excel file bytes
        """
        try:
            self.create_workbook('Quality Report')

            # Headers
            headers = ['Inspection ID', 'Product', 'Date', 'Inspector', 'Defects Found', 'Quality Score', 'Result', 'Notes']
            for col, header in enumerate(headers, 1):
                self.sheet.cell(row=1, column=col, value=header)

            self._apply_header_style(1, len(headers))

            # Data rows
            for row, inspection in enumerate(inspections, 2):
                self.sheet.cell(row=row, column=1, value=inspection.get('id'))
                self.sheet.cell(row=row, column=2, value=inspection.get('product_name'))
                self.sheet.cell(row=row, column=3, value=inspection.get('date'))
                self.sheet.cell(row=row, column=4, value=inspection.get('inspector'))
                self.sheet.cell(row=row, column=5, value=inspection.get('defects_found', 0))

                # Quality score
                score = inspection.get('quality_score', 0)
                score_cell = self.sheet.cell(row=row, column=6, value=score)
                if score >= 90:
                    score_cell.fill = PatternFill(start_color=self.COLORS['success'], fill_type='solid')
                elif score >= 70:
                    score_cell.fill = PatternFill(start_color=self.COLORS['warning'], fill_type='solid')
                else:
                    score_cell.fill = PatternFill(start_color=self.COLORS['danger'], fill_type='solid')

                # Result
                result = inspection.get('result', 'FAIL')
                result_cell = self.sheet.cell(row=row, column=7, value=result)
                if result == 'PASS':
                    result_cell.fill = PatternFill(start_color=self.COLORS['success'], fill_type='solid')
                else:
                    result_cell.fill = PatternFill(start_color=self.COLORS['danger'], fill_type='solid')

                self.sheet.cell(row=row, column=8, value=inspection.get('notes'))

            # Auto-adjust
            self._auto_adjust_column_width()

            # Summary stats if provided
            if summary_stats:
                summary_row = len(inspections) + 3
                self.sheet.cell(row=summary_row, column=1, value='Summary Statistics').font = Font(bold=True, size=14)

                summary_row += 1
                self.sheet.cell(row=summary_row, column=1, value='Total Inspections:')
                self.sheet.cell(row=summary_row, column=2, value=summary_stats.get('total', 0))

                summary_row += 1
                self.sheet.cell(row=summary_row, column=1, value='Pass Rate:')
                pass_rate_cell = self.sheet.cell(row=summary_row, column=2, value=summary_stats.get('pass_rate', 0))
                pass_rate_cell.number_format = '0.00%'

                summary_row += 1
                self.sheet.cell(row=summary_row, column=1, value='Average Quality Score:')
                self.sheet.cell(row=summary_row, column=2, value=summary_stats.get('avg_score', 0))

            # Save to bytes
            output = io.BytesIO()
            self.workbook.save(output)
            output.seek(0)

            logger.info(f'Exported {len(inspections)} quality inspections to Excel')
            return output.getvalue()

        except Exception as e:
            logger.error(f'Quality report export failed: {e}')
            raise


# Singleton instance
_excel_service = None


def get_excel_export_service() -> ExcelExportService:
    """Get Excel export service singleton"""
    global _excel_service
    if _excel_service is None:
        _excel_service = ExcelExportService()
    return _excel_service
