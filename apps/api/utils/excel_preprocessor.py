"""
Excel Preprocessor Utility

Handles common Excel parsing challenges:
- Merged cells detection and unmerging
- Multi-level header flattening
- Empty row/column removal
- Data normalization
"""

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class ExcelPreprocessor:
    """Preprocesses Excel files for easier parsing by handling merged cells and complex structures."""

    def __init__(self, excel_path: str | Path):
        """
        Initialize preprocessor with Excel file path.

        Args:
            excel_path: Path to Excel file
        """
        self.excel_path = Path(excel_path)
        self.workbook = None
        self.processed_sheets = {}

    def __enter__(self):
        """Context manager entry - load workbook."""
        self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit - close workbook."""
        if self.workbook:
            self.workbook.close()

    def unmerge_and_fill_cells(self, sheet_name: str = None) -> Worksheet:
        """
        Unmerge all cells in a sheet and fill with the merged value.

        Args:
            sheet_name: Name of sheet to process. If None, uses active sheet.

        Returns:
            Processed worksheet
        """
        if not self.workbook:
            raise ValueError("Workbook not loaded. Use context manager or call load()")

        ws = self.workbook[sheet_name] if sheet_name else self.workbook.active

        # Get all merged cell ranges
        merged_ranges = list(ws.merged_cells.ranges)

        logger.info(f"Found {len(merged_ranges)} merged cell ranges in sheet '{ws.title}'")

        # Process each merged range
        for merged_range in merged_ranges:
            # Get the value from the top-left cell
            min_col, min_row, max_col, max_row = merged_range.bounds
            merged_value = ws.cell(min_row, min_col).value

            # Unmerge the cells
            ws.unmerge_cells(str(merged_range))

            # Fill all cells in the range with the value
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row, col).value = merged_value

        logger.info(f"Unmerged and filled {len(merged_ranges)} cell ranges")
        return ws

    def detect_header_rows(self, sheet: Worksheet, max_header_rows: int = 5) -> int:
        """
        Detect number of header rows by finding first row with mostly non-None values.

        Args:
            sheet: Worksheet to analyze
            max_header_rows: Maximum rows to check for headers

        Returns:
            Number of header rows detected
        """
        for row_idx in range(1, max_header_rows + 1):
            row_values = [sheet.cell(row_idx, col).value for col in range(1, sheet.max_column + 1)]
            non_none_count = sum(1 for v in row_values if v is not None)

            # If >50% of cells have values, likely a header
            if non_none_count > len(row_values) * 0.5:
                logger.info(f"Detected header row at row {row_idx}")
                return row_idx

        return 1  # Default to first row

    def sheet_to_dataframe(
        self,
        sheet_name: str = None,
        unmerge: bool = True,
        skip_empty_rows: bool = True,
        skip_empty_cols: bool = True,
    ) -> pd.DataFrame:
        """
        Convert Excel sheet to pandas DataFrame with preprocessing.

        Args:
            sheet_name: Sheet to process. If None, uses active sheet.
            unmerge: Whether to unmerge cells before conversion
            skip_empty_rows: Skip rows that are completely empty
            skip_empty_cols: Skip columns that are completely empty

        Returns:
            Preprocessed DataFrame
        """
        if not self.workbook:
            raise ValueError("Workbook not loaded. Use context manager or call load()")

        # Get or process sheet
        if unmerge:
            ws = self.unmerge_and_fill_cells(sheet_name)
        else:
            ws = self.workbook[sheet_name] if sheet_name else self.workbook.active

        # Convert to list of lists
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))

        # Create DataFrame
        df = pd.DataFrame(data)

        # Skip empty rows
        if skip_empty_rows:
            df = df.dropna(how="all")

        # Skip empty columns
        if skip_empty_cols:
            df = df.dropna(axis=1, how="all")

        # Reset index
        df = df.reset_index(drop=True)

        logger.info(f"Converted sheet '{ws.title}' to DataFrame: {df.shape}")
        return df

    def analyze_structure(self, sheet_name: str = None) -> Dict[str, Any]:
        """
        Analyze Excel sheet structure for debugging and understanding.

        Args:
            sheet_name: Sheet to analyze

        Returns:
            Dictionary with structure information
        """
        if not self.workbook:
            raise ValueError("Workbook not loaded")

        ws = self.workbook[sheet_name] if sheet_name else self.workbook.active

        # Count merged cells
        merged_count = len(list(ws.merged_cells.ranges))

        # Detect empty rows/cols
        empty_rows = []
        for row_idx in range(1, ws.max_row + 1):
            if all(ws.cell(row_idx, col).value is None for col in range(1, ws.max_column + 1)):
                empty_rows.append(row_idx)

        empty_cols = []
        for col_idx in range(1, ws.max_column + 1):
            if all(ws.cell(row, col_idx).value is None for row in range(1, ws.max_row + 1)):
                empty_cols.append(col_idx)

        # Sample first few rows
        sample_rows = []
        for row_idx in range(1, min(6, ws.max_row + 1)):
            row_data = {}
            for col_idx in range(1, min(11, ws.max_column + 1)):
                value = ws.cell(row_idx, col_idx).value
                if value is not None:
                    row_data[f"Col{col_idx}"] = str(value)[:50]
            if row_data:
                sample_rows.append({f"Row{row_idx}": row_data})

        return {
            "sheet_name": ws.title,
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "merged_cells_count": merged_count,
            "empty_rows_count": len(empty_rows),
            "empty_cols_count": len(empty_cols),
            "sample_data": sample_rows,
        }

    def save_preprocessed(
        self, output_path: str | Path, sheet_name: str = None, format: str = "csv"
    ):
        """
        Save preprocessed Excel data to file.

        Args:
            output_path: Where to save the file
            sheet_name: Sheet to process
            format: Output format ('csv', 'json', 'excel')
        """
        df = self.sheet_to_dataframe(sheet_name)
        output_path = Path(output_path)

        if format == "csv":
            df.to_csv(output_path, index=False, encoding="utf-8-sig")
        elif format == "json":
            df.to_json(output_path, orient="records", force_ascii=False, indent=2)
        elif format == "excel":
            df.to_excel(output_path, index=False, engine="openpyxl")
        else:
            raise ValueError(f"Unsupported format: {format}")

        logger.info(f"Saved preprocessed data to {output_path}")


def quick_preview(
    excel_path: str | Path, sheet_name: str = None, max_rows: int = 10
) -> pd.DataFrame:
    """
    Quick preview of Excel file with automatic preprocessing.

    Args:
        excel_path: Path to Excel file
        sheet_name: Sheet to preview
        max_rows: Maximum rows to show

    Returns:
        Preview DataFrame
    """
    with ExcelPreprocessor(excel_path) as preprocessor:
        df = preprocessor.sheet_to_dataframe(sheet_name)
        return df.head(max_rows)


def analyze_excel_file(excel_path: str | Path) -> Dict[str, Any]:
    """
    Comprehensive analysis of Excel file structure.

    Args:
        excel_path: Path to Excel file

    Returns:
        Analysis results for all sheets
    """
    with ExcelPreprocessor(excel_path) as preprocessor:
        analysis = {
            "file_path": str(excel_path),
            "total_sheets": len(preprocessor.workbook.sheetnames),
            "sheet_names": preprocessor.workbook.sheetnames,
            "sheets": {},
        }

        for sheet_name in preprocessor.workbook.sheetnames:
            analysis["sheets"][sheet_name] = preprocessor.analyze_structure(sheet_name)

        return analysis
