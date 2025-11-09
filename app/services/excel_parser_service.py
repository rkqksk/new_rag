"""
Excel Parser Service
공식 Excel 파일을 파싱하여 크롤링 데이터와 비교
"""

import json
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


@dataclass
class ExcelProduct:
    """Excel에서 추출한 제품 정보"""

    code: str
    spec: str
    packaging: str
    mold: str
    cost: Optional[float]
    price: Optional[float]
    production: Optional[str]
    note: str
    images: List[str]  # 이미지 파일 경로
    row_number: int


class ExcelParserService:
    """Excel 파일 파싱 및 비교 서비스"""

    def __init__(self, upload_dir: str = "data/excel_uploads"):
        self.upload_dir = Path(upload_dir)
        self.raw_dir = self.upload_dir / "raw"
        self.processed_dir = self.upload_dir / "processed"
        self.images_dir = self.upload_dir / "images"

        # 폴더 확인
        for dir_path in [self.raw_dir, self.processed_dir, self.images_dir]:
            dir_path.mkdir(parents=True, exist_ok=True)

        logger.info(f"✅ ExcelParserService initialized: {self.upload_dir}")

    def parse_excel(self, filename: str) -> List[ExcelProduct]:
        """
        Excel 파일 파싱 (다중 컬럼 레이아웃 지원)

        Args:
            filename: Excel 파일명 (raw/ 폴더 내)

        Returns:
            List[ExcelProduct]: 파싱된 제품 목록
        """
        try:
            import openpyxl
            from openpyxl.drawing.image import Image as XLImage
        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            return []

        excel_path = self.raw_dir / filename
        if not excel_path.exists():
            logger.error(f"Excel file not found: {excel_path}")
            return []

        logger.info(f"📖 Parsing Excel: {filename}")

        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            sheet = workbook.active

            products = []

            # 다중 컬럼 레이아웃 감지: Row 5에 "CODE"가 여러 개 있는지 확인
            row5_values = [cell.value for cell in sheet[5]]
            code_columns = [idx for idx, val in enumerate(row5_values, start=1) if val == "CODE"]

            if len(code_columns) > 1:
                logger.info(
                    f"Detected multi-column layout with {len(code_columns)} products per row group"
                )
                products = self._parse_multi_column_layout(sheet, code_columns)
            else:
                # 기존 단순 테이블 레이아웃
                logger.info("Using simple table layout parser")
                products = self._parse_simple_table_layout(sheet)

            # 이미지 추출 (별도 처리)
            self._extract_images(sheet, filename)

            logger.info(f"✅ Parsed {len(products)} products from Excel")
            return products

        except Exception as e:
            logger.error(f"Error parsing Excel: {e}", exc_info=True)
            return []

    def _parse_multi_column_layout(self, sheet, code_columns: list) -> List[ExcelProduct]:
        """다중 컬럼 레이아웃 파싱 - 전체 시트 스캔하여 모든 제품 추출

        Layout pattern repeats throughout sheet:
        - Row N: "CODE" labels
        - Row N+1: Product codes (col+1)
        - Row N+2: SPEC labels
        - Row N+3: Specs (col+1)
        """
        products = []
        max_row = sheet.max_row

        logger.info(f"Scanning sheet up to row {max_row} for product groups...")

        # Scan every 10 rows looking for "CODE" pattern (product groups repeat)
        for start_row in range(5, max_row, 10):
            # Check if this row has CODE labels
            row_values = [cell.value for cell in sheet[start_row]]
            if "CODE" not in row_values:
                continue

            # Find CODE columns in this row
            local_code_cols = [idx for idx, val in enumerate(row_values, start=1) if val == "CODE"]

            if not local_code_cols:
                continue

            logger.debug(
                f"Found product group at row {start_row} with {len(local_code_cols)} columns"
            )

            # Extract products from this group
            for col_idx in local_code_cols:
                try:
                    # Product code is in col_idx + 1, row start_row+1
                    code_cell = sheet.cell(row=start_row + 1, column=col_idx + 1)
                    code_value = code_cell.value

                    if not code_value or str(code_value).startswith("="):
                        continue

                    # Spec is in col_idx + 1, row start_row+2 (after CODE+1)
                    spec_cell = sheet.cell(row=start_row + 2, column=col_idx + 1)
                    spec_value = spec_cell.value or ""

                    # If empty, try next row (dimensions)
                    if not spec_value or str(spec_value).strip() == "":
                        spec_cell = sheet.cell(row=start_row + 3, column=col_idx + 1)
                        spec_value = spec_cell.value or ""

                    # Packaging info
                    packaging_cell = sheet.cell(row=start_row + 5, column=col_idx + 2)
                    packaging_value = packaging_cell.value or ""

                    # Mold info
                    mold_cell = sheet.cell(row=start_row + 6, column=col_idx + 2)
                    mold_value = mold_cell.value or ""

                    product = ExcelProduct(
                        code=str(code_value).strip(),
                        spec=str(spec_value).strip(),
                        packaging=str(packaging_value),
                        mold=str(mold_value),
                        cost=None,
                        price=None,
                        production="",
                        note="",
                        images=[],
                        row_number=start_row + 1,
                    )

                    products.append(product)

                except Exception as e:
                    logger.debug(f"Skipping col {col_idx} at row {start_row}: {e}")
                    continue

        logger.info(f"✅ Extracted {len(products)} total products from full sheet")
        return products

    def _parse_simple_table_layout(self, sheet) -> List[ExcelProduct]:
        """단순 테이블 레이아웃 파싱 (한 행에 한 제품)"""
        products = []

        # Find header row
        headers = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                headers[cell.value] = col_idx

        logger.info(f"Found columns: {list(headers.keys())}")

        # Parse data rows (from row 2)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            code_col = headers.get("Code") or headers.get("code") or headers.get("CODE")
            if not code_col:
                continue

            code_value = row[code_col - 1].value
            if not code_value:
                continue

            product = ExcelProduct(
                code=str(code_value),
                spec=str(row[headers.get("Spec", 1) - 1].value or ""),
                packaging=str(row[headers.get("포장", 2) - 1].value or ""),
                mold=str(row[headers.get("금형", 3) - 1].value or ""),
                cost=self._parse_number(row[headers.get("원가", 4) - 1].value),
                price=self._parse_number(row[headers.get("판매", 5) - 1].value),
                production=str(row[headers.get("생산량", 6) - 1].value or ""),
                note=str(row[headers.get("비고", 7) - 1].value or ""),
                images=[],
                row_number=row_idx,
            )

            products.append(product)

        return products

    def _parse_number(self, value: Any) -> Optional[float]:
        """숫자 값 파싱"""
        try:
            if value is None:
                return None
            return float(value)
        except (ValueError, TypeError):
            return None

    def _extract_images(self, sheet, filename: str):
        """
        Excel에서 이미지 추출

        Args:
            sheet: openpyxl worksheet
            filename: Excel 파일명
        """
        try:
            if not hasattr(sheet, "_images"):
                logger.info("No images found in Excel")
                return

            image_count = 0
            for image in sheet._images:
                # 이미지 저장
                image_filename = f"{Path(filename).stem}_img_{image_count}.png"
                image_path = self.images_dir / image_filename

                with open(image_path, "wb") as f:
                    f.write(image._data())

                image_count += 1
                logger.info(f"📷 Extracted image: {image_filename}")

            logger.info(f"✅ Extracted {image_count} images")

        except Exception as e:
            logger.error(f"Error extracting images: {e}")

    def compare_with_database(
        self, excel_products: List[ExcelProduct], db_products: List[Dict[str, Any]]
    ) -> Dict[str, Any]:
        """
        Excel 데이터와 DB 데이터 비교

        Args:
            excel_products: Excel에서 파싱한 제품
            db_products: Qdrant에서 가져온 제품

        Returns:
            비교 리포트
        """
        # Excel 제품 코드 맵
        excel_map = {p.code: p for p in excel_products}

        # DB 제품 코드 맵
        db_map = {}
        for p in db_products:
            specs = p.get("specifications", {})
            code = specs.get("제품 코드", "N/A")
            if code and code != "N/A":
                db_map[code] = p

        # 비교
        missing_in_db = []
        missing_codes_in_db = []
        spec_mismatches = []

        for code, excel_prod in excel_map.items():
            if code not in db_map:
                missing_in_db.append(excel_prod)
            else:
                # 스펙 비교
                db_prod = db_map[code]
                db_specs = db_prod.get("specifications", {})
                db_spec = db_specs.get("사양", "")

                if excel_prod.spec != db_spec:
                    spec_mismatches.append(
                        {"code": code, "excel_spec": excel_prod.spec, "db_spec": db_spec}
                    )

        # DB에만 있는 제품 (제품 코드 없는 것)
        for p in db_products:
            specs = p.get("specifications", {})
            code = specs.get("제품 코드", "N/A")
            if code == "N/A":
                missing_codes_in_db.append(p.get("product_name", "Unknown"))

        report = {
            "total_excel": len(excel_products),
            "total_db": len(db_products),
            "total_db_with_codes": len(db_map),
            "missing_in_db": len(missing_in_db),
            "missing_codes_in_db": len(missing_codes_in_db),
            "spec_mismatches": len(spec_mismatches),
            "details": {
                "missing_in_db": [p.code for p in missing_in_db],
                "missing_codes": missing_codes_in_db,
                "spec_mismatches": spec_mismatches,
            },
        }

        logger.info(f"📊 Comparison Report: {report}")
        return report

    def save_parsed_data(self, filename: str, products: List[ExcelProduct]):
        """파싱된 데이터를 JSON으로 저장"""
        output_file = self.processed_dir / f"{Path(filename).stem}.json"

        data = [
            {
                "code": p.code,
                "spec": p.spec,
                "packaging": p.packaging,
                "mold": p.mold,
                "cost": p.cost,
                "price": p.price,
                "production": p.production,
                "note": p.note,
                "images": p.images,
                "row_number": p.row_number,
            }
            for p in products
        ]

        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        logger.info(f"💾 Saved parsed data: {output_file}")
        return str(output_file)


if __name__ == "__main__":
    # 테스트
    service = ExcelParserService()
    print(f"✅ Service initialized")
    print(f"📁 Upload your Excel files to: {service.raw_dir}")
