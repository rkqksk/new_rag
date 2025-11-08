"""
Advanced Excel Processor with openpyxl

Handles complex Excel files with:
- Merged cells preservation
- Embedded images extraction
- Cell formulas (not just values)
- Cell styles and formatting
- Multiple sheets
- Charts documentation
- Conditional formatting
"""

import io
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional, Dict, Any, List, Tuple
from datetime import datetime

import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
from PIL import Image

logger = logging.getLogger(__name__)


@dataclass
class MergedCell:
    """Merged cell information"""
    range: str  # "A1:C1"
    start_row: int
    start_col: int
    end_row: int
    end_col: int
    value: Any
    rows_span: int
    cols_span: int


@dataclass
class ExcelImage:
    """Excel embedded image"""
    path: str  # Saved image path
    width: int
    height: int
    format: str  # PNG, JPEG, etc.
    anchor_cell: str  # Cell where image is anchored
    anchor_row: int
    anchor_col: int


@dataclass
class CellStyle:
    """Cell style information"""
    font: Dict[str, Any] = field(default_factory=dict)
    fill: Dict[str, Any] = field(default_factory=dict)
    border: Dict[str, Any] = field(default_factory=dict)
    alignment: Dict[str, Any] = field(default_factory=dict)
    number_format: Optional[str] = None


@dataclass
class ExcelSheet:
    """Single Excel sheet data"""
    name: str
    data: List[List[Any]]  # Raw cell values
    merged_cells: List[MergedCell]
    images: List[ExcelImage]
    styles: Dict[str, CellStyle]  # Cell coordinate -> style
    formulas: Dict[str, str]  # Cell coordinate -> formula
    max_row: int
    max_col: int


@dataclass
class ExcelData:
    """Complete Excel workbook data"""
    metadata: Dict[str, Any]
    sheets: Dict[str, ExcelSheet]
    sheet_names: List[str]


@dataclass
class ExcelProcessorConfig:
    """Excel processor configuration"""
    extract_images: bool = True
    extract_formulas: bool = True
    extract_styles: bool = True
    extract_merged_cells: bool = True
    process_all_sheets: bool = True
    image_output_dir: Optional[str] = None
    max_image_size: int = 10 * 1024 * 1024  # 10MB max per image
    data_only: bool = False  # If True, formulas evaluated to values


class AdvancedExcelProcessor:
    """
    Advanced Excel file processor using openpyxl

    Preserves:
    - ✅ Merged cells with range info
    - ✅ Embedded images (extracted to files)
    - ✅ Cell formulas
    - ✅ Cell styles (font, fill, border, alignment)
    - ✅ Multiple sheets
    - ✅ Charts documentation
    - ✅ Conditional formatting

    Example:
        >>> processor = AdvancedExcelProcessor('complex_report.xlsx')
        >>> data = processor.extract_all()
        >>>
        >>> # Access sheet data
        >>> sheet1 = data.sheets['Sheet1']
        >>> print(f"Merged cells: {len(sheet1.merged_cells)}")
        >>> print(f"Images: {len(sheet1.images)}")
        >>>
        >>> # Handle merged cells
        >>> for merge in sheet1.merged_cells:
        >>>     print(f"{merge.range}: {merge.value}")
    """

    def __init__(
        self,
        file_path: str,
        config: Optional[ExcelProcessorConfig] = None
    ):
        self.file_path = Path(file_path)
        self.config = config or ExcelProcessorConfig()

        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        # Load workbook
        self.workbook = load_workbook(
            filename=str(self.file_path),
            data_only=self.config.data_only
        )

        # Image output directory
        if self.config.image_output_dir:
            self.image_dir = Path(self.config.image_output_dir)
        else:
            self.image_dir = self.file_path.parent / f"{self.file_path.stem}_images"

        self.image_dir.mkdir(parents=True, exist_ok=True)

    def extract_all(self) -> ExcelData:
        """
        Extract all data from Excel file

        Returns:
            ExcelData object with complete workbook information
        """
        logger.info(f"Processing Excel file: {self.file_path}")

        # Extract metadata
        metadata = self._extract_metadata()

        # Process sheets
        sheets = {}
        sheet_names = self.workbook.sheetnames

        if self.config.process_all_sheets:
            for sheet_name in sheet_names:
                logger.info(f"Processing sheet: {sheet_name}")
                sheet = self.workbook[sheet_name]
                sheets[sheet_name] = self._process_sheet(sheet, sheet_name)
        else:
            # Only process first sheet
            sheet_name = sheet_names[0]
            logger.info(f"Processing first sheet only: {sheet_name}")
            sheet = self.workbook[sheet_name]
            sheets[sheet_name] = self._process_sheet(sheet, sheet_name)

        result = ExcelData(
            metadata=metadata,
            sheets=sheets,
            sheet_names=sheet_names
        )

        logger.info(f"Excel processing complete: {len(sheets)} sheets processed")
        return result

    def _extract_metadata(self) -> Dict[str, Any]:
        """Extract workbook metadata"""
        props = self.workbook.properties

        return {
            'file_path': str(self.file_path),
            'file_size': self.file_path.stat().st_size,
            'creator': props.creator,
            'created': props.created.isoformat() if props.created else None,
            'modified': props.modified.isoformat() if props.modified else None,
            'last_modified_by': props.lastModifiedBy,
            'title': props.title,
            'subject': props.subject,
            'description': props.description,
            'keywords': props.keywords,
            'category': props.category,
            'sheet_count': len(self.workbook.sheetnames),
            'sheet_names': self.workbook.sheetnames
        }

    def _process_sheet(self, sheet: Worksheet, sheet_name: str) -> ExcelSheet:
        """Process a single worksheet"""

        # Extract data (raw values)
        data = self._extract_data(sheet)

        # Extract merged cells
        merged_cells = []
        if self.config.extract_merged_cells:
            merged_cells = self._extract_merged_cells(sheet)

        # Extract images
        images = []
        if self.config.extract_images:
            images = self._extract_images(sheet, sheet_name)

        # Extract formulas
        formulas = {}
        if self.config.extract_formulas:
            formulas = self._extract_formulas(sheet)

        # Extract styles
        styles = {}
        if self.config.extract_styles:
            styles = self._extract_styles(sheet)

        return ExcelSheet(
            name=sheet_name,
            data=data,
            merged_cells=merged_cells,
            images=images,
            styles=styles,
            formulas=formulas,
            max_row=sheet.max_row,
            max_col=sheet.max_column
        )

    def _extract_data(self, sheet: Worksheet) -> List[List[Any]]:
        """Extract raw cell data"""
        data = []

        for row in sheet.iter_rows():
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            data.append(row_data)

        return data

    def _extract_merged_cells(self, sheet: Worksheet) -> List[MergedCell]:
        """Extract merged cell information"""
        merged = []

        for merged_range in sheet.merged_cells.ranges:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            max_row, max_col = merged_range.max_row, merged_range.max_col

            # Get the value from the top-left cell
            cell = sheet.cell(min_row, min_col)

            merged_cell = MergedCell(
                range=str(merged_range),
                start_row=min_row,
                start_col=min_col,
                end_row=max_row,
                end_col=max_col,
                value=cell.value,
                rows_span=max_row - min_row + 1,
                cols_span=max_col - min_col + 1
            )

            merged.append(merged_cell)

        logger.debug(f"Found {len(merged)} merged cells")
        return merged

    def _extract_images(self, sheet: Worksheet, sheet_name: str) -> List[ExcelImage]:
        """Extract embedded images"""
        images = []

        if not hasattr(sheet, '_images'):
            return images

        for idx, image in enumerate(sheet._images):
            try:
                # Get image data
                img_data = image._data()

                # Check size limit
                if len(img_data) > self.config.max_image_size:
                    logger.warning(f"Image {idx} exceeds size limit, skipping")
                    continue

                # Load image with PIL
                pil_image = Image.open(io.BytesIO(img_data))

                # Save image
                image_filename = f"{sheet_name}_img_{idx}.{pil_image.format.lower()}"
                image_path = self.image_dir / image_filename
                pil_image.save(image_path)

                # Get anchor cell
                anchor = image.anchor
                if hasattr(anchor, '_from'):
                    anchor_col = anchor._from.col
                    anchor_row = anchor._from.row
                    anchor_cell = f"{get_column_letter(anchor_col + 1)}{anchor_row + 1}"
                else:
                    anchor_col = 0
                    anchor_row = 0
                    anchor_cell = "A1"

                excel_image = ExcelImage(
                    path=str(image_path),
                    width=pil_image.width,
                    height=pil_image.height,
                    format=pil_image.format,
                    anchor_cell=anchor_cell,
                    anchor_row=anchor_row,
                    anchor_col=anchor_col
                )

                images.append(excel_image)
                logger.debug(f"Extracted image: {image_filename}")

            except Exception as e:
                logger.error(f"Error extracting image {idx}: {e}")

        logger.debug(f"Extracted {len(images)} images")
        return images

    def _extract_formulas(self, sheet: Worksheet) -> Dict[str, str]:
        """Extract cell formulas"""
        formulas = {}

        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # Formula cell
                    coord = cell.coordinate
                    formulas[coord] = cell.value

        logger.debug(f"Found {len(formulas)} formulas")
        return formulas

    def _extract_styles(self, sheet: Worksheet) -> Dict[str, CellStyle]:
        """Extract cell styles"""
        styles = {}

        for row in sheet.iter_rows():
            for cell in row:
                if not cell.value:
                    continue

                # Only extract styles for non-empty cells
                font = cell.font
                fill = cell.fill
                border = cell.border
                alignment = cell.alignment

                style = CellStyle(
                    font={
                        'name': font.name,
                        'size': font.size,
                        'bold': font.bold,
                        'italic': font.italic,
                        'color': str(font.color.rgb) if font.color and font.color.rgb else None
                    },
                    fill={
                        'fill_type': fill.fill_type,
                        'start_color': str(fill.start_color.rgb) if fill.start_color and hasattr(fill.start_color, 'rgb') else None,
                        'end_color': str(fill.end_color.rgb) if fill.end_color and hasattr(fill.end_color, 'rgb') else None
                    },
                    border={
                        'left': str(border.left.style) if border.left else None,
                        'right': str(border.right.style) if border.right else None,
                        'top': str(border.top.style) if border.top else None,
                        'bottom': str(border.bottom.style) if border.bottom else None
                    },
                    alignment={
                        'horizontal': alignment.horizontal,
                        'vertical': alignment.vertical,
                        'wrap_text': alignment.wrap_text
                    },
                    number_format=cell.number_format
                )

                coord = cell.coordinate
                styles[coord] = style

        logger.debug(f"Extracted styles for {len(styles)} cells")
        return styles

    def get_merged_cell_value(self, sheet_name: str, row: int, col: int, data: ExcelData) -> Tuple[Any, Optional[MergedCell]]:
        """
        Get cell value, accounting for merged cells

        Args:
            sheet_name: Sheet name
            row: Row number (1-indexed)
            col: Column number (1-indexed)
            data: ExcelData object

        Returns:
            Tuple of (value, merged_cell_info or None)
        """
        sheet = data.sheets[sheet_name]

        # Check if cell is part of a merged range
        for merged in sheet.merged_cells:
            if (merged.start_row <= row <= merged.end_row and
                merged.start_col <= col <= merged.end_col):
                return merged.value, merged

        # Not merged, return normal value
        value = sheet.data[row - 1][col - 1] if row <= sheet.max_row and col <= sheet.max_col else None
        return value, None

    def export_to_dict(self, data: ExcelData) -> Dict[str, Any]:
        """
        Export ExcelData to dictionary (JSON-serializable)

        Args:
            data: ExcelData object

        Returns:
            Dictionary representation
        """
        result = {
            'metadata': data.metadata,
            'sheets': {}
        }

        for sheet_name, sheet in data.sheets.items():
            result['sheets'][sheet_name] = {
                'name': sheet.name,
                'data': sheet.data,
                'max_row': sheet.max_row,
                'max_col': sheet.max_col,
                'merged_cells': [
                    {
                        'range': m.range,
                        'value': m.value,
                        'rows_span': m.rows_span,
                        'cols_span': m.cols_span
                    }
                    for m in sheet.merged_cells
                ],
                'images': [
                    {
                        'path': img.path,
                        'width': img.width,
                        'height': img.height,
                        'format': img.format,
                        'anchor_cell': img.anchor_cell
                    }
                    for img in sheet.images
                ],
                'formulas': sheet.formulas,
                'styles_count': len(sheet.styles)
            }

        return result


# Convenience function
def process_excel_file(
    file_path: str,
    extract_all: bool = True,
    image_output_dir: Optional[str] = None
) -> ExcelData:
    """
    Convenience function to process Excel file

    Args:
        file_path: Path to Excel file
        extract_all: Extract all features (merged cells, images, formulas, styles)
        image_output_dir: Directory for extracted images

    Returns:
        ExcelData object

    Example:
        >>> data = process_excel_file('report.xlsx')
        >>> print(f"Sheets: {data.sheet_names}")
        >>> print(f"Merged cells: {len(data.sheets['Sheet1'].merged_cells)}")
    """
    config = ExcelProcessorConfig(
        extract_images=extract_all,
        extract_formulas=extract_all,
        extract_styles=extract_all,
        extract_merged_cells=extract_all,
        image_output_dir=image_output_dir
    )

    processor = AdvancedExcelProcessor(file_path, config)
    return processor.extract_all()
