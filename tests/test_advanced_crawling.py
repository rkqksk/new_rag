"""
Integration tests for Advanced Crawling & Data Acquisition

Tests:
- Static crawler
- Dynamic crawler (Playwright)
- Multi-strategy crawler with auto-detection
- Advanced Excel processor
- Authentication manager
- Session manager
- Anti-bot evasion
"""

import asyncio
import pytest
from pathlib import Path
import tempfile

# Import crawling services
from app.services.crawling.static_crawler import StaticCrawler, StaticCrawlerConfig
from app.services.crawling.dynamic_crawler import DynamicCrawler, PlaywrightConfig
from app.services.crawling.multi_strategy_crawler import (
    MultiStrategyCrawler,
    CrawlMethod,
    CrawlConfig
)
from app.services.crawling.auth_manager import (
    AuthenticationManager,
    AuthType,
    AuthCredentials
)
from app.services.crawling.session_manager import SessionManager
from app.services.crawling.evasion import (
    AntiDetectionManager,
    EvasionConfig,
    RateLimiter
)

# Import data processing services
from app.services.data_processing.excel_processor import (
    AdvancedExcelProcessor,
    ExcelProcessorConfig,
    process_excel_file
)


# ==================== Static Crawler Tests ====================

@pytest.mark.asyncio
async def test_static_crawler_basic():
    """Test basic static HTML crawling"""
    crawler = StaticCrawler()

    # Test with a simple HTTP endpoint
    result = await crawler.crawl('http://httpbin.org/html')

    assert result is not None
    assert result['status_code'] == 200
    assert 'content' in result
    assert result['url'] == 'http://httpbin.org/html'


@pytest.mark.asyncio
async def test_static_crawler_json():
    """Test JSON API endpoint"""
    crawler = StaticCrawler()

    result = await crawler.crawl(
        'http://httpbin.org/json',
        parse_html=False
    )

    assert result is not None
    assert result['status_code'] == 200
    assert 'text' in result


@pytest.mark.asyncio
async def test_static_crawler_retry():
    """Test retry logic"""
    config = StaticCrawlerConfig(
        timeout=5,
        max_retries=2,
        retry_delay=0.5
    )
    crawler = StaticCrawler(config)

    # This should succeed even with short timeout
    result = await crawler.crawl('http://httpbin.org/delay/1')

    assert result is not None
    assert result['status_code'] == 200


@pytest.mark.asyncio
async def test_static_crawler_extract_links():
    """Test link extraction"""
    crawler = StaticCrawler()

    links = await crawler.extract_links('http://httpbin.org/links/5')

    assert links is not None
    assert len(links) > 0


# ==================== Dynamic Crawler Tests ====================

@pytest.mark.asyncio
@pytest.mark.skip(reason="Requires Playwright installation")
async def test_dynamic_crawler_basic():
    """Test dynamic crawling with Playwright"""
    config = PlaywrightConfig(headless=True, stealth_mode=True)
    crawler = DynamicCrawler(config)

    try:
        result = await crawler.crawl('http://httpbin.org/html')

        assert result is not None
        assert 'content' in result
        assert 'title' in result
        assert result['url'] == 'http://httpbin.org/html'

    finally:
        await crawler.close()


@pytest.mark.asyncio
@pytest.mark.skip(reason="Requires Playwright installation")
async def test_dynamic_crawler_javascript():
    """Test JavaScript rendering"""
    config = PlaywrightConfig(headless=True)

    async with DynamicCrawler(config) as crawler:
        # Test with a page that requires JavaScript
        result = await crawler.crawl(
            'http://httpbin.org/html',
            wait_for_timeout=1000
        )

        assert result is not None
        assert len(result['content']) > 0


# ==================== Multi-Strategy Crawler Tests ====================

@pytest.mark.asyncio
async def test_multi_strategy_auto_detect():
    """Test auto-detection of crawling method"""
    crawler = MultiStrategyCrawler()

    # Should detect as static HTML
    method = await crawler.detect_method('http://httpbin.org/html')

    assert method in [CrawlMethod.STATIC, CrawlMethod.DYNAMIC]


@pytest.mark.asyncio
async def test_multi_strategy_crawl_static():
    """Test multi-strategy crawler with static method"""
    async with MultiStrategyCrawler() as crawler:
        result = await crawler.crawl(
            'http://httpbin.org/html',
            method=CrawlMethod.STATIC
        )

        assert result is not None
        assert result['crawl_method'] == 'static'
        assert 'content' in result


@pytest.mark.asyncio
async def test_multi_strategy_stats():
    """Test crawler statistics"""
    async with MultiStrategyCrawler() as crawler:
        # Make a request
        await crawler.crawl('http://httpbin.org/html')

        stats = crawler.get_stats()

        assert stats['total_requests'] == 1
        assert stats['static_requests'] >= 0
        assert 'rate_limit' in stats


# ==================== Authentication Manager Tests ====================

@pytest.mark.asyncio
async def test_auth_manager_basic():
    """Test basic authentication"""
    auth_manager = AuthenticationManager()

    credentials = AuthCredentials(
        username='test_user',
        password='test_pass'
    )

    client = await auth_manager.authenticate(
        AuthType.BASIC,
        credentials
    )

    assert client is not None
    # Client should have auth configured
    assert client.auth is not None


@pytest.mark.asyncio
async def test_auth_manager_api_key():
    """Test API key authentication"""
    auth_manager = AuthenticationManager()

    credentials = AuthCredentials(
        api_key='test-api-key-12345',
        api_key_header='X-API-Key'
    )

    client = await auth_manager.authenticate(
        AuthType.API_KEY,
        credentials
    )

    assert client is not None
    assert 'X-API-Key' in client.headers
    assert client.headers['X-API-Key'] == 'test-api-key-12345'


def test_auth_manager_totp_generation():
    """Test TOTP secret generation"""
    secret = AuthenticationManager.generate_totp_secret()

    assert secret is not None
    assert len(secret) == 32  # Base32 encoded
    assert secret.isalnum()


def test_auth_manager_totp_uri():
    """Test TOTP provisioning URI"""
    secret = 'JBSWY3DPEHPK3PXP'

    uri = AuthenticationManager.get_totp_provisioning_uri(
        secret=secret,
        name='test@example.com',
        issuer='TestApp'
    )

    assert uri.startswith('otpauth://totp/')
    assert 'TestApp' in uri
    assert 'test@example.com' in uri


# ==================== Session Manager Tests ====================

def test_session_manager_save_load():
    """Test session save and load"""
    import httpx

    with tempfile.TemporaryDirectory() as tmpdir:
        manager = SessionManager(storage_dir=tmpdir)

        # Create a client with cookies
        client = httpx.AsyncClient()
        client.cookies.set('test_cookie', 'test_value')

        # Save session
        manager.save_session('test_session', client)

        # Load session
        loaded_client = manager.load_session('test_session')

        assert loaded_client is not None
        assert 'test_cookie' in loaded_client.cookies


def test_session_manager_list_sessions():
    """Test listing sessions"""
    with tempfile.TemporaryDirectory() as tmpdir:
        manager = SessionManager(storage_dir=tmpdir)

        # Create some sessions
        import httpx
        for i in range(3):
            client = httpx.AsyncClient()
            manager.save_session(f'session_{i}', client)

        sessions = manager.list_sessions()

        assert len(sessions) == 3
        assert 'session_0' in sessions
        assert 'session_1' in sessions
        assert 'session_2' in sessions


def test_session_manager_delete():
    """Test session deletion"""
    with tempfile.TemporaryDirectory() as tmpdir:
        manager = SessionManager(storage_dir=tmpdir)

        import httpx
        client = httpx.AsyncClient()
        manager.save_session('test_session', client)

        # Verify exists
        loaded = manager.load_session('test_session')
        assert loaded is not None

        # Delete
        manager.delete_session('test_session')

        # Verify deleted
        loaded = manager.load_session('test_session')
        assert loaded is None


# ==================== Anti-Detection Tests ====================

def test_evasion_manager_user_agent():
    """Test user agent rotation"""
    config = EvasionConfig(rotate_user_agent=True)
    manager = AntiDetectionManager(config)

    ua1 = manager.get_random_user_agent()
    ua2 = manager.get_random_user_agent()

    assert ua1 is not None
    assert ua2 is not None
    # They might be the same, but both should be valid


def test_evasion_manager_headers():
    """Test header generation"""
    config = EvasionConfig(
        rotate_user_agent=True,
        randomize_headers=True,
        spoof_referrer=True
    )
    manager = AntiDetectionManager(config)

    headers = manager.get_headers(url='https://example.com')

    assert 'User-Agent' in headers
    assert 'Accept' in headers
    assert 'Accept-Language' in headers
    assert 'Referer' in headers


def test_evasion_manager_proxy():
    """Test proxy rotation"""
    config = EvasionConfig(
        use_proxies=True,
        proxy_list=['http://proxy1.com:8080', 'http://proxy2.com:8080'],
        proxy_rotation_interval=2
    )
    manager = AntiDetectionManager(config)

    proxy1 = manager.get_proxy()
    proxy2 = manager.get_proxy()  # Should be same (not rotated yet)

    assert proxy1 is not None
    assert proxy2 is not None


@pytest.mark.asyncio
async def test_evasion_manager_delay():
    """Test human-like delays"""
    import time

    config = EvasionConfig(
        min_delay=0.1,
        max_delay=0.2,
        randomize_delay=True
    )
    manager = AntiDetectionManager(config)

    start = time.time()
    await manager.human_delay()
    elapsed = time.time() - start

    # Should be between min and max delay
    assert 0.1 <= elapsed <= 0.3  # Allow small margin


@pytest.mark.asyncio
async def test_rate_limiter():
    """Test rate limiter"""
    limiter = RateLimiter(max_requests=5, time_window=1.0)

    # Make 5 requests quickly
    for _ in range(5):
        await limiter.acquire()

    # 6th request should be delayed
    import time
    start = time.time()
    await limiter.acquire()
    elapsed = time.time() - start

    # Should have waited
    assert elapsed > 0.5


# ==================== Excel Processor Tests ====================

def test_excel_processor_basic():
    """Test basic Excel processing"""
    # Create a simple test Excel file
    import openpyxl

    with tempfile.TemporaryDirectory() as tmpdir:
        test_file = Path(tmpdir) / 'test.xlsx'

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'TestSheet'

        # Add data
        ws['A1'] = 'Name'
        ws['B1'] = 'Age'
        ws['A2'] = 'John'
        ws['B2'] = 30

        wb.save(test_file)

        # Process file
        processor = AdvancedExcelProcessor(str(test_file))
        data = processor.extract_all()

        assert data is not None
        assert 'TestSheet' in data.sheets
        assert data.sheets['TestSheet'].max_row == 2
        assert data.sheets['TestSheet'].max_col == 2


def test_excel_processor_merged_cells():
    """Test merged cell handling"""
    import openpyxl

    with tempfile.TemporaryDirectory() as tmpdir:
        test_file = Path(tmpdir) / 'test_merged.xlsx'

        # Create workbook with merged cells
        wb = openpyxl.Workbook()
        ws = wb.active

        # Merge cells
        ws.merge_cells('A1:C1')
        ws['A1'] = 'Merged Header'

        ws['A2'] = 'Data1'
        ws['B2'] = 'Data2'
        ws['C2'] = 'Data3'

        wb.save(test_file)

        # Process file
        processor = AdvancedExcelProcessor(
            str(test_file),
            ExcelProcessorConfig(extract_merged_cells=True)
        )
        data = processor.extract_all()

        sheet = data.sheets[ws.title]

        assert len(sheet.merged_cells) == 1
        merged = sheet.merged_cells[0]
        assert merged.range == 'A1:C1'
        assert merged.value == 'Merged Header'
        assert merged.cols_span == 3
        assert merged.rows_span == 1


def test_excel_processor_formulas():
    """Test formula preservation"""
    import openpyxl

    with tempfile.TemporaryDirectory() as tmpdir:
        test_file = Path(tmpdir) / 'test_formulas.xlsx'

        # Create workbook with formulas
        wb = openpyxl.Workbook()
        ws = wb.active

        ws['A1'] = 10
        ws['A2'] = 20
        ws['A3'] = '=SUM(A1:A2)'

        wb.save(test_file)

        # Process file (data_only=False to preserve formulas)
        processor = AdvancedExcelProcessor(
            str(test_file),
            ExcelProcessorConfig(
                extract_formulas=True,
                data_only=False
            )
        )
        data = processor.extract_all()

        sheet = data.sheets[ws.title]

        assert 'A3' in sheet.formulas
        assert sheet.formulas['A3'] == '=SUM(A1:A2)'


def test_excel_processor_convenience_function():
    """Test convenience function"""
    import openpyxl

    with tempfile.TemporaryDirectory() as tmpdir:
        test_file = Path(tmpdir) / 'test_convenience.xlsx'

        # Create simple workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Test'
        wb.save(test_file)

        # Process using convenience function
        data = process_excel_file(str(test_file))

        assert data is not None
        assert len(data.sheets) > 0


# ==================== Integration Tests ====================

@pytest.mark.asyncio
async def test_full_crawl_pipeline():
    """Test complete crawling pipeline"""
    config = CrawlConfig(
        method=CrawlMethod.STATIC,
        use_evasion=True,
        rate_limit_requests=5,
        rate_limit_window=10.0
    )

    async with MultiStrategyCrawler(config) as crawler:
        # Crawl a simple page
        result = await crawler.crawl('http://httpbin.org/html')

        assert result is not None
        assert 'content' in result
        assert result['crawl_method'] == 'static'

        # Check stats
        stats = crawler.get_stats()
        assert stats['total_requests'] == 1


if __name__ == '__main__':
    # Run tests
    pytest.main([__file__, '-v'])
