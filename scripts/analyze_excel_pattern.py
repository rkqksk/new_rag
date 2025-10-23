#!/usr/bin/env python3
"""엑셀 구조 패턴 분석"""

import openpyxl
from pathlib import Path

excel_file = Path("data/excel_uploads/raw/제품 리스트_1.PE.xlsx")
wb = openpyxl.load_workbook(excel_file, data_only=True)
sheet = wb.active

print("=" * 80)
print("엑셀 구조 패턴 분석")
print("=" * 80)

# 이미지 위치 확인
print(f"\n총 행: {sheet.max_row}, 총 열: {sheet.max_column}")

# 이미지가 있는지 확인
if hasattr(sheet, '_images'):
    print(f"\n이미지 개수: {len(sheet._images)}")
    
    # 처음 5개 이미지 위치 분석
    for idx, img in enumerate(sheet._images[:5], 1):
        if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
            img_row = img.anchor._from.row + 1  # 0-based to 1-based
            img_col = img.anchor._from.col + 1
            print(f"\n이미지 {idx}: Row {img_row}, Col {img_col}")
            
            # 해당 이미지 아래 15행의 데이터 확인
            print(f"  이미지 아래 데이터:")
            for offset in range(0, 15):
                check_row = img_row + offset
                if check_row <= sheet.max_row:
                    # 이미지와 같은 열 및 오른쪽 열들 확인
                    row_data = []
                    for col in range(img_col, min(img_col + 3, sheet.max_column + 1)):
                        cell_value = sheet.cell(row=check_row, column=col).value
                        if cell_value:
                            row_data.append(f"Col{col}={cell_value}")
                    
                    if row_data:
                        print(f"    Row {check_row} [+{offset}]: {', '.join(row_data)}")

# CODE 셀 패턴 분석
print("\n" + "=" * 80)
print("CODE 셀 패턴 분석 (첫 10개)")
print("=" * 80)

code_cells = []
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
    for cell in row:
        if cell.value == "CODE":
            code_cells.append(cell)

for idx, code_cell in enumerate(code_cells[:10], 1):
    row_idx = code_cell.row
    col_idx = code_cell.column
    
    print(f"\nCODE #{idx}: {code_cell.coordinate} (Row {row_idx}, Col {col_idx})")
    
    # CODE 아래 12행 확인
    print("  아래 12행:")
    for offset in range(1, 13):
        check_row = row_idx + offset
        cell_val = sheet.cell(row=check_row, column=col_idx).value
        next_col_val = sheet.cell(row=check_row, column=col_idx + 1).value
        
        print(f"    Row {check_row} [+{offset}]: {code_cell.coordinate[0]}{check_row}={cell_val}, 옆={next_col_val}")

# 다음 CODE까지의 거리 확인
print("\n" + "=" * 80)
print("CODE 간 거리 패턴")
print("=" * 80)

for idx in range(min(9, len(code_cells) - 1)):
    curr_row = code_cells[idx].row
    next_row = code_cells[idx + 1].row
    distance = next_row - curr_row
    print(f"CODE {idx+1} (Row {curr_row}) → CODE {idx+2} (Row {next_row}): {distance}행 간격")

wb.close()
