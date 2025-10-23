"""
Intelligent Price List Parser

Automatically detects merged cell patterns and parses Excel price lists
with minimal manual intervention.
"""

import openpyxl
import json
import re
from pathlib import Path
from typing import List, Dict, Any, Tuple
from collections import defaultdict


class IntelligentPriceParser:
    """Parses price list Excel files by analyzing merged cell patterns."""

    def __init__(self, excel_path: str | Path):
        self.excel_path = Path(excel_path)
        self.workbook = None
        self.column_mappings = {}  # sheet_name -> {col_idx: category_info}

    def __enter__(self):
        self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.workbook:
            self.workbook.close()

    def analyze_category_structure(self, sheet_name: str) -> Dict[int, Dict[str, Any]]:
        """
        Analyze merged cells in row 4 to determine column-to-category mapping.

        Returns:
            Dictionary mapping column index to category info:
            {col_idx: {'category': 'name', 'sub_header': 'header'}}
        """
        ws = self.workbook[sheet_name]
        merged_ranges = list(ws.merged_cells.ranges)

        # Find row 4 category headers (merged cells)
        row4_categories = []
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds

            if min_row == 4 and max_row == 4:
                category_name = ws.cell(min_row, min_col).value
                if category_name:
                    row4_categories.append({
                        'category': str(category_name).strip(),
                        'min_col': min_col,
                        'max_col': max_col
                    })

        # Build column mapping
        col_mapping = {}

        for cat_info in row4_categories:
            category = cat_info['category']
            min_col = cat_info['min_col']
            max_col = cat_info['max_col']

            # Map each column in this range to the category
            for col_idx in range(min_col, max_col + 1):
                # Get sub-header from row 5
                sub_header = ws.cell(5, col_idx).value
                sub_header = str(sub_header).strip() if sub_header else None

                col_mapping[col_idx] = {
                    'category': category,
                    'sub_header': sub_header
                }

        # Handle non-merged columns in row 4
        for col_idx in range(1, ws.max_column + 1):
            if col_idx not in col_mapping:
                header = ws.cell(4, col_idx).value
                if header:
                    # This is a standalone category
                    col_mapping[col_idx] = {
                        'category': str(header).strip(),
                        'sub_header': ws.cell(5, col_idx).value
                    }

        return col_mapping

    def parse_price(self, price_str: Any) -> Tuple[int, int]:
        """Extract regular and discount price from string like '140 (100)'."""
        if not price_str or price_str == '-':
            return None, None

        price_str = str(price_str).strip()

        # Match pattern: "140 (100)" or just "140"
        match = re.match(r'(\d+)\s*(?:\((\d+)\))?', price_str)
        if match:
            regular_price = int(match.group(1))
            discount_price = int(match.group(2)) if match.group(2) else regular_price
            return regular_price, discount_price

        return None, None

    def parse_container_sheet(self, ws, col_mapping: Dict[int, Dict]) -> List[Dict[str, Any]]:
        """Parse '용기' sheet using intelligent column mapping."""
        records = []

        # Find capacity column (should be in column with sub_header '용량')
        capacity_col = None
        for col_idx, info in col_mapping.items():
            if info.get('sub_header') == '용량' and info['category'] in ['브로우', '거품 브로우']:
                capacity_col = col_idx
                break

        if not capacity_col:
            capacity_col = 2  # Default

        # Parse data rows (starting from row 6)
        for row_idx in range(6, ws.max_row + 1):
            capacity_cell = ws.cell(row_idx, capacity_col).value
            if not capacity_cell:
                continue

            capacity = str(capacity_cell).strip()

            # Process each column
            for col_idx, col_info in col_mapping.items():
                category = col_info['category']
                sub_header = col_info['sub_header']

                # Skip capacity column
                if col_idx == capacity_col:
                    continue

                price_cell = ws.cell(row_idx, col_idx).value
                if price_cell and price_cell != '-':
                    regular, discount = self.parse_price(price_cell)
                    if regular:
                        record = {
                            'category': '용기',
                            'sub_category': category,
                            'capacity_ml': capacity,
                            'regular_price': regular,
                            'discount_price': discount,
                            'unit': '원'
                        }

                        # Add material or other info based on sub_header
                        if category in ['브로우', '거품 브로우']:
                            if sub_header and sub_header != '용량':
                                record['material'] = sub_header
                        elif category == '인쇄':
                            if sub_header:
                                record['material'] = sub_header
                        elif category == '라벨':
                            record['label_type'] = sub_header if sub_header else '라벨'

                        records.append(record)

        return records

    def parse_cap_pump_sheet(self, ws, col_mapping: Dict[int, Dict]) -> List[Dict[str, Any]]:
        """Parse '캡,펌프' sheet using intelligent column mapping."""
        records = []

        # This sheet has more complex structure with merged cells in multiple rows
        # We'll focus on extracting price data based on column categories

        for row_idx in range(6, ws.max_row + 1):
            # Get product type from column 2 (if merged vertically)
            product_type_cell = ws.cell(row_idx, 2).value

            for col_idx, col_info in col_mapping.items():
                if col_idx <= 2:  # Skip first two columns
                    continue

                category = col_info['category']
                sub_header = col_info['sub_header']

                price_cell = ws.cell(row_idx, col_idx).value
                if price_cell:
                    regular, discount = self.parse_price(price_cell)
                    if regular:
                        records.append({
                            'category': '캡,펌프',
                            'sub_category': category,
                            'product_type': str(product_type_cell).strip() if product_type_cell else '',
                            'size': sub_header if sub_header else '',
                            'regular_price': regular,
                            'discount_price': discount,
                            'unit': '원'
                        })

        return records

    def parse_coating_sheet(self, ws, col_mapping: Dict[int, Dict]) -> List[Dict[str, Any]]:
        """Parse '코팅' sheet using intelligent column mapping."""
        records = []

        # Find capacity column
        capacity_col = 2

        for row_idx in range(6, ws.max_row + 1):
            capacity_cell = ws.cell(row_idx, capacity_col).value
            if not capacity_cell:
                continue

            capacity = str(capacity_cell).strip()

            for col_idx, col_info in col_mapping.items():
                if col_idx == capacity_col:
                    continue

                category = col_info['category']
                sub_header = col_info['sub_header']

                price_cell = ws.cell(row_idx, col_idx).value
                if price_cell and price_cell != '-':
                    regular, discount = self.parse_price(price_cell)
                    if regular:
                        records.append({
                            'category': category,
                            'capacity_range': capacity,
                            'coating_type': sub_header if sub_header else category,
                            'regular_price': regular,
                            'discount_price': discount,
                            'unit': '원'
                        })

        return records

    def parse_all_sheets(self) -> List[Dict[str, Any]]:
        """Parse all sheets in the workbook."""
        all_records = []

        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]

            print(f"\n📋 파싱 중: '{sheet_name}'")

            # Analyze structure
            col_mapping = self.analyze_category_structure(sheet_name)
            self.column_mappings[sheet_name] = col_mapping

            print(f"  🔍 감지된 카테고리: {set(info['category'] for info in col_mapping.values())}")

            # Parse based on sheet type
            if sheet_name == '용기':
                records = self.parse_container_sheet(ws, col_mapping)
            elif sheet_name == '캡,펌프':
                records = self.parse_cap_pump_sheet(ws, col_mapping)
            elif sheet_name == '코팅':
                records = self.parse_coating_sheet(ws, col_mapping)
            else:
                records = []

            print(f"  ✅ 추출: {len(records)}개 레코드")
            all_records.extend(records)

        return all_records

    def generate_structure_report(self) -> Dict[str, Any]:
        """Generate detailed report of detected structure."""
        report = {
            'file': str(self.excel_path),
            'sheets': {}
        }

        for sheet_name, col_mapping in self.column_mappings.items():
            # Group by category
            categories = defaultdict(list)
            for col_idx, info in col_mapping.items():
                categories[info['category']].append({
                    'column': col_idx,
                    'column_letter': openpyxl.utils.get_column_letter(col_idx),
                    'sub_header': info['sub_header']
                })

            report['sheets'][sheet_name] = {
                'total_columns': len(col_mapping),
                'categories': {
                    cat: {
                        'columns': sorted([c['column'] for c in cols]),
                        'column_range': f"{cols[0]['column_letter']}:{cols[-1]['column_letter']}",
                        'sub_headers': [c['sub_header'] for c in cols if c['sub_header']]
                    }
                    for cat, cols in categories.items()
                }
            }

        return report


def main():
    """Main execution."""
    excel_path = Path("data/excel_uploads/price_list/2025 매출 단가표 (1).xlsx")

    print("=" * 80)
    print("🤖 Intelligent Price List Parser")
    print("=" * 80)

    with IntelligentPriceParser(excel_path) as parser:
        # Parse all sheets
        records = parser.parse_all_sheets()

        # Generate structure report
        structure_report = parser.generate_structure_report()

        print(f"\n{'=' * 80}")
        print(f"✅ 총 {len(records)}개 레코드 추출")
        print("=" * 80)

        # Save results
        output_json = Path("data/excel_uploads/price_list/intelligent_parsed.json")
        with open(output_json, 'w', encoding='utf-8') as f:
            json.dump(records, f, ensure_ascii=False, indent=2)
        print(f"\n💾 데이터 저장: {output_json}")

        # Save structure report
        report_path = Path("data/excel_uploads/price_list/structure_report.json")
        with open(report_path, 'w', encoding='utf-8') as f:
            json.dump(structure_report, f, ensure_ascii=False, indent=2)
        print(f"📊 구조 분석 보고서: {report_path}")

        # Show structure
        print(f"\n📋 감지된 구조:")
        for sheet_name, sheet_info in structure_report['sheets'].items():
            print(f"\n  [{sheet_name}]")
            for category, cat_info in sheet_info['categories'].items():
                print(f"    - {category}: 열 {cat_info['column_range']} → {cat_info['sub_headers'][:3]}...")

        # Sample records
        print(f"\n📊 샘플 레코드:")
        for i, record in enumerate(records[:10]):
            print(f"  {i+1}. {record}")


if __name__ == "__main__":
    main()
