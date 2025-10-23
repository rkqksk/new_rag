#!/usr/bin/env python3
"""
Universal Document Parser
범용 문서 파서 - PaddleOCR (무료) + OpenAI Vision (백업) 지원
"""

import sys
import json
import os
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
import io

# UTF-8 강제 설정
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

try:
    import openpyxl
    from PIL import Image, ImageDraw, ImageFont
    import numpy as np
except ImportError as e:
    print(f"❌ Missing dependencies: {e}")
    sys.exit(1)

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)


class BaseVisionProvider:
    """Vision Provider 기본 클래스"""

    def analyze_image(self, image: Image.Image, material: str) -> Dict[str, Any]:
        """이미지 분석 - 하위 클래스에서 구현"""
        raise NotImplementedError


class PaddleOCRProvider(BaseVisionProvider):
    """PaddleOCR 기반 Provider (무료)"""

    def __init__(self):
        from paddleocr import PaddleOCR
        logger.info("🚀 Initializing PaddleOCR Provider...")
        self.ocr = PaddleOCR(lang='korean')

        # Claude Vision에서 학습한 한글 라벨
        self.korean_labels = {
            '포장': 'packaging',
            '금형': 'mold',
            '원가': 'cost',
            '판매': 'price',
            '생산량': 'production',
            '비고': 'note',
            'CODE': 'product_code',
            'SPEC': 'spec'
        }
        logger.info("✅ PaddleOCR Provider initialized")

    def analyze_image(self, image: Image.Image, material: str) -> Dict[str, Any]:
        """PaddleOCR로 이미지 분석"""
        logger.info("🔍 Analyzing with PaddleOCR...")

        # OCR 실행
        img_array = np.array(image)
        ocr_result = self.ocr.predict(img_array)

        # 텍스트 추출
        texts = self._extract_texts(ocr_result)

        # 한글 라벨 찾기
        labels = self._find_korean_labels(texts)

        # 제품 데이터 추출
        products = self._extract_products(texts, labels, material)

        return {
            'products': products,
            'pattern_analysis': f'PaddleOCR extracted {len(labels)} labels, {len(products)} products',
            'method': 'paddleocr'
        }

    def _extract_texts(self, ocr_result) -> List[Dict[str, Any]]:
        """OCR 결과에서 텍스트 추출"""
        texts = []
        rec_texts = ocr_result.get('rec_texts', [])
        rec_scores = ocr_result.get('rec_scores', [])
        dt_polys = ocr_result.get('dt_polys', [])

        for idx in range(len(rec_texts)):
            text = rec_texts[idx]
            confidence = rec_scores[idx] if idx < len(rec_scores) else 1.0
            bbox = dt_polys[idx] if idx < len(dt_polys) else []

            if not bbox:
                continue

            center_x = sum([float(p[0]) for p in bbox]) / len(bbox)
            center_y = sum([float(p[1]) for p in bbox]) / len(bbox)

            texts.append({
                'text': str(text).strip(),
                'center': (center_x, center_y),
                'confidence': float(confidence)
            })

        logger.info(f"   ✅ OCR: {len(texts)} text elements")
        return texts

    def _find_korean_labels(self, texts: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """한글 라벨 위치 찾기"""
        labels = []
        for item in texts:
            text = item['text']
            for korean_label, field_name in self.korean_labels.items():
                if korean_label in text:
                    labels.append({
                        'label': korean_label,
                        'field': field_name,
                        'center': item['center']
                    })
                    break

        logger.info(f"   📌 Found {len(labels)} Korean labels")
        return labels

    def _extract_products(
        self,
        texts: List[Dict[str, Any]],
        labels: List[Dict[str, Any]],
        material: str
    ) -> List[Dict[str, Any]]:
        """제품 데이터 추출"""
        products = []

        # 라벨을 세로 위치로 그룹화 (같은 제품)
        sorted_labels = sorted(labels, key=lambda x: x['center'][1])

        current_group = []
        for i, label in enumerate(sorted_labels):
            if not current_group:
                current_group.append(label)
            else:
                prev_y = current_group[-1]['center'][1]
                curr_y = label['center'][1]

                # 세로 거리 200 이상이면 새 제품
                if curr_y - prev_y > 200:
                    products.append(self._extract_product_data(texts, current_group, material))
                    current_group = [label]
                else:
                    current_group.append(label)

        # 마지막 그룹
        if current_group:
            products.append(self._extract_product_data(texts, current_group, material))

        logger.info(f"   ✅ Extracted {len(products)} products")
        return products

    def _extract_product_data(
        self,
        texts: List[Dict[str, Any]],
        label_group: List[Dict[str, Any]],
        material: str
    ) -> Dict[str, Any]:
        """라벨 그룹에서 제품 데이터 추출"""
        product = {'material': material}

        for label_info in label_group:
            field = label_info['field']
            value = self._find_value_near_label(texts, label_info['center'])
            product[field] = self._infer_type(value)

        return product

    def _find_value_near_label(
        self,
        texts: List[Dict[str, Any]],
        label_center: tuple,
        max_distance: float = 200
    ) -> Optional[str]:
        """라벨 근처의 값 찾기 (오른쪽)"""
        label_x, label_y = label_center

        candidates = []
        for item in texts:
            text = item['text']
            center_x, center_y = item['center']

            # 오른쪽이 아니면 스킵
            if center_x <= label_x:
                continue

            x_dist = center_x - label_x
            y_dist = abs(center_y - label_y)

            if x_dist > max_distance or y_dist > 30:
                continue

            # 라벨이면 스킵
            if any(label in text for label in self.korean_labels.keys()):
                continue

            candidates.append({
                'text': text,
                'distance': x_dist + y_dist
            })

        if candidates:
            closest = min(candidates, key=lambda x: x['distance'])
            return closest['text']

        return None

    def _infer_type(self, value: Optional[str]) -> Any:
        """타입 추론"""
        if value is None or value.strip() == '':
            return None

        value = value.strip()

        try:
            if '.' in value:
                return float(value)
            return int(value)
        except ValueError:
            return value


class OpenAIVisionProvider(BaseVisionProvider):
    """OpenAI Vision API Provider (백업)"""

    def __init__(self):
        import openai
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            raise ValueError("OPENAI_API_KEY not set")

        logger.info("🚀 Initializing OpenAI Vision Provider...")
        self.client = openai.OpenAI(api_key=api_key)
        logger.info("✅ OpenAI Vision Provider initialized")

    def analyze_image(self, image: Image.Image, material: str) -> Dict[str, Any]:
        """OpenAI Vision으로 이미지 분석"""
        logger.info("🔍 Analyzing with OpenAI Vision...")

        import base64
        from io import BytesIO

        # 이미지를 base64로 변환
        buffered = BytesIO()
        image.save(buffered, format="PNG")
        image_b64 = base64.b64encode(buffered.getvalue()).decode('utf-8')

        prompt = f"""Analyze this Excel product catalog image. This is a {material} product list.

Extract ALL product information as JSON array:

```json
[
  {{
    "product_code": "code from CODE field",
    "spec": "material / capacity",
    "dimensions": "size if present",
    "packaging_label": "포장 or packaging",
    "packaging_value": "packaging data",
    "mold_label": "금형 or mold",
    "mold_value": "mold data",
    "cost_label": "원가 or cost",
    "cost_value": cost as number or null,
    "price_label": "판매 or price",
    "price_value": price as number or null,
    "production_label": "생산량 or production",
    "production_value": "production data",
    "note_label": "비고 or note",
    "note_value": "note data",
    "material": "{material}"
  }}
]
```

Return ONLY valid JSON."""

        try:
            response = self.client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/png;base64,{image_b64}"
                                }
                            },
                            {
                                "type": "text",
                                "text": prompt
                            }
                        ]
                    }
                ],
                max_tokens=4096
            )

            response_text = response.choices[0].message.content
            logger.info("✅ Received response from OpenAI Vision")

            # JSON 추출
            import re
            json_match = re.search(r'```json\s*(.*?)\s*```', response_text, re.DOTALL)
            if not json_match:
                json_match = re.search(r'\[.*\]', response_text, re.DOTALL)

            if json_match:
                products_json = json_match.group(1) if json_match.lastindex else json_match.group()
                products = json.loads(products_json)

                return {
                    'products': products,
                    'pattern_analysis': f'OpenAI Vision extracted {len(products)} products',
                    'method': 'openai_vision'
                }
            else:
                logger.error("❌ No JSON in response")
                return {'products': [], 'pattern_analysis': '', 'method': 'openai_vision'}

        except Exception as e:
            logger.error(f"❌ OpenAI Vision error: {e}")
            return {'products': [], 'error': str(e), 'method': 'openai_vision'}


class UniversalDocumentParser:
    """
    범용 문서 파서

    Provider 선택:
    - VISION_PROVIDER=paddleocr (기본, 무료)
    - VISION_PROVIDER=openai (백업, OpenAI 크레딧 사용)
    """

    def __init__(self, provider: str = None):
        """
        Args:
            provider: 'paddleocr' or 'openai'. None이면 환경변수 사용
        """
        if provider is None:
            provider = os.getenv('VISION_PROVIDER', 'paddleocr')

        logger.info(f"🔧 Initializing Universal Document Parser (provider={provider})")

        if provider == 'paddleocr':
            self.vision_provider = PaddleOCRProvider()
        elif provider == 'openai':
            self.vision_provider = OpenAIVisionProvider()
        else:
            raise ValueError(f"Unknown provider: {provider}. Use 'paddleocr' or 'openai'")

        self.provider_name = provider
        self.output_dir = Path("data/excel_uploads/universal_parsed")
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def extract_images(self, excel_path: Path, sheet_name: str = None) -> List[Dict[str, Any]]:
        """
        openpyxl로 엑셀에 embedded된 이미지 추출

        모든 provider에서 공통으로 사용
        """
        logger.info("   📷 Extracting embedded images from Excel...")

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb[sheet_name] if sheet_name else wb.active

        images_data = []

        if not hasattr(sheet, '_images') or not sheet._images:
            logger.info("   ℹ️  No images found in sheet")
            wb.close()
            return images_data

        logger.info(f"   Found {len(sheet._images)} embedded images")

        # 이미지 저장 디렉토리
        image_dir = self.output_dir / excel_path.stem / "images"
        image_dir.mkdir(parents=True, exist_ok=True)

        for idx, img in enumerate(sheet._images, 1):
            try:
                # 이미지 위치 정보
                if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                    img_row = img.anchor._from.row + 1  # 0-based to 1-based
                    img_col = img.anchor._from.col + 1
                else:
                    img_row, img_col = 0, 0

                # 이미지 파일 경로
                image_path = image_dir / f"image_{idx:03d}_r{img_row}c{img_col}.png"

                # PIL Image로 변환 및 저장
                if hasattr(img, '_data'):
                    from io import BytesIO
                    pil_img = Image.open(BytesIO(img._data()))
                    pil_img.save(image_path)

                    images_data.append({
                        'index': idx,
                        'row': img_row,
                        'column': img_col,
                        'path': str(image_path),
                        'size': {'width': pil_img.width, 'height': pil_img.height}
                    })

            except Exception as e:
                logger.warning(f"   ⚠️  Failed to extract image {idx}: {e}")

        wb.close()

        logger.info(f"   ✅ Extracted {len(images_data)} images to {image_dir.name}/")
        return images_data

    def excel_to_image(
        self,
        excel_path: Path,
        start_row: int = 1,
        end_row: int = 50
    ) -> Image.Image:
        """Excel을 이미지로 변환 (한글 폰트 지원)"""
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb.active

        # 자동으로 최대 컬럼 감지
        end_col = sheet.max_column
        logger.info(f"   Auto-detected columns: 1-{end_col} ({openpyxl.utils.get_column_letter(end_col)})")

        # 컬럼/행 크기 계산
        col_widths = {}
        for col in range(1, min(end_col + 1, sheet.max_column + 1)):
            col_letter = openpyxl.utils.get_column_letter(col)
            width = sheet.column_dimensions[col_letter].width or 10
            col_widths[col] = max(width, 8) * 10

        row_heights = {}
        for row in range(start_row, min(end_row + 1, sheet.max_row + 1)):
            height = sheet.row_dimensions[row].height or 15
            row_heights[row] = max(height, 15) * 1.5

        img_width = int(sum(col_widths.values()))
        img_height = int(sum(row_heights.values()))

        img = Image.new('RGB', (img_width, img_height), 'white')
        draw = ImageDraw.Draw(img)

        # 한글 폰트
        korean_fonts = [
            "/System/Library/Fonts/AppleSDGothicNeo.ttc",
            "/System/Library/Fonts/AppleGothic.ttf",
            "/System/Library/Fonts/Supplemental/Arial Unicode.ttf"
        ]

        font = None
        for font_path in korean_fonts:
            try:
                font = ImageFont.truetype(font_path, 14)
                break
            except:
                continue

        if not font:
            font = ImageFont.load_default()

        # 셀 그리기
        y_offset = 0
        for row in range(start_row, min(end_row + 1, sheet.max_row + 1)):
            x_offset = 0
            row_height = row_heights.get(row, 20)

            for col in range(1, min(end_col + 1, sheet.max_column + 1)):
                cell = sheet.cell(row=row, column=col)
                col_width = col_widths.get(col, 80)

                draw.rectangle(
                    [x_offset, y_offset, x_offset + col_width, y_offset + row_height],
                    outline='gray',
                    width=1
                )

                if cell.value:
                    text = str(cell.value)[:100]
                    draw.text(
                        (x_offset + 5, y_offset + 5),
                        text,
                        fill='black',
                        font=font
                    )

                x_offset += col_width

            y_offset += row_height

        wb.close()
        return img

    def parse_excel(
        self,
        excel_path: Path,
        rows_per_batch: int = 30,
        num_batches: int = 3
    ) -> Dict[str, Any]:
        """Excel 파일 파싱"""
        logger.info(f"\n📊 Parsing: {excel_path.name} (provider={self.provider_name})")

        # 재질 감지
        filename_upper = excel_path.name.upper()
        if 'PETG' in filename_upper:
            material = 'PETG'
        elif 'PET' in filename_upper:
            material = 'PET'
        elif 'PE' in filename_upper:
            material = 'PE'
        elif 'PP' in filename_upper:
            material = 'PP'
        else:
            material = 'Other'

        logger.info(f"   Material: {material}")

        # 1단계: openpyxl로 이미지 추출 (provider 무관)
        logger.info("\n🖼️  Step 1: Extract embedded images")
        images = self.extract_images(excel_path)

        all_products = []

        # 2단계: Vision API로 텍스트 데이터 추출
        logger.info(f"\n📝 Step 2: Extract text data with {self.provider_name}")

        # 배치 처리
        for batch_idx in range(num_batches):
            start_row = 1 + (batch_idx * rows_per_batch)
            end_row = start_row + rows_per_batch - 1

            logger.info(f"\n   Batch {batch_idx + 1}/{num_batches}: rows {start_row}-{end_row}")

            # 이미지 생성
            image = self.excel_to_image(excel_path, start_row, end_row)

            # 이미지 저장
            image_path = self.output_dir / f"{excel_path.stem}_batch{batch_idx + 1}.png"
            image.save(image_path)
            logger.info(f"   💾 Image: {image_path.name}")

            # Vision 분석
            result = self.vision_provider.analyze_image(image, material)

            if result['products']:
                all_products.extend(result['products'])
                logger.info(f"   ✅ Batch {batch_idx + 1}: {len(result['products'])} products")

        # 결과 저장
        output = {
            'file': excel_path.name,
            'material': material,
            'provider': self.provider_name,
            'total_products': len(all_products),
            'total_images': len(images),
            'products': all_products,
            'images': images
        }

        output_file = self.output_dir / f"{excel_path.stem}_parsed.json"
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(output, f, ensure_ascii=False, indent=2)

        logger.info(f"\n💾 Saved: {output_file}")
        logger.info(f"✅ Total: {len(all_products)} products")

        return output


def main():
    """테스트"""

    # 환경변수 또는 명령줄 인자로 provider 선택
    provider = os.getenv('VISION_PROVIDER', 'paddleocr')

    if len(sys.argv) > 1:
        provider = sys.argv[1]

    print("=" * 80)
    print(f"Universal Document Parser (Provider: {provider})")
    print("=" * 80)

    parser = UniversalDocumentParser(provider=provider)

    excel_file = Path("data/excel_uploads/raw/제품 리스트_1.PE.xlsx")

    if not excel_file.exists():
        print(f"❌ File not found: {excel_file}")
        return

    result = parser.parse_excel(excel_file, rows_per_batch=30, num_batches=3)

    print(f"\n{'='*80}")
    print("Results:")
    print(f"{'='*80}")
    print(f"  Provider: {result['provider']}")
    print(f"  Material: {result['material']}")
    print(f"  Products: {result['total_products']}")
    print(f"  Images: {result['total_images']}")
    print(f"{'='*80}")

    # 샘플 출력
    if result['products']:
        print(f"\n📝 Sample (first 3):")
        for i, p in enumerate(result['products'][:3], 1):
            print(f"\n{i}. {p.get('product_code', 'N/A')}")
            print(f"   Spec: {p.get('spec', 'N/A')}")
            print(f"   포장: {p.get('packaging_value', 'N/A')}")
            print(f"   원가: {p.get('cost_value', 'N/A')}")
            print(f"   판매: {p.get('price_value', 'N/A')}")


if __name__ == "__main__":
    main()
