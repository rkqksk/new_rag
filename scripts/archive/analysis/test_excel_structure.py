#!/usr/bin/env python3
"""Test Excel structure analysis"""

import openpyxl
from pathlib import Path

excel_file = Path("data/excel_uploads/raw/제품 리스트_1.PE.xlsx")

wb = openpyxl.load_workbook(excel_file, data_only=True)
sheet = wb.active

print(f"Sheet dimensions: {sheet.dimensions}")
print(f"Max row: {sheet.max_row}, Max col: {sheet.max_column}")
print()

# Find all CODE cells
print("=" * 80)
print("Searching for 'CODE' cells and analyzing structure...")
print("=" * 80)

code_cells = []
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
    for cell in row:
        if cell.value == "CODE":
            code_cells.append(cell)

print(f"\nFound {len(code_cells)} 'CODE' cells\n")

# Analyze first 5 CODE cells
for i, code_cell in enumerate(code_cells[:5], 1):
    print(f"\n{'='*60}")
    print(f"CODE Cell #{i}: {code_cell.coordinate}")
    print(f"{'='*60}")

    row_idx = code_cell.row
    col_idx = code_cell.column

    # Check surrounding cells (vertical pattern)
    print(f"\nVertical pattern (same column {col_idx}):")
    for offset in range(-2, 10):
        check_row = row_idx + offset
        cell = sheet.cell(row=check_row, column=col_idx)
        print(f"  Row {check_row} [{offset:+3}]: {cell.coordinate:6} = {repr(cell.value)[:50]}")

    # Check horizontal pattern
    print(f"\nHorizontal pattern (same row {row_idx}):")
    for offset in range(-1, 3):
        check_col = col_idx + offset
        cell = sheet.cell(row=row_idx, column=check_col)
        print(f"  Col {check_col} [{offset:+2}]: {cell.coordinate:6} = {repr(cell.value)[:50]}")

    # Check for images in cells above
    print(f"\nChecking for images above CODE cell...")
    has_image = False
    for img_offset in range(-10, 0):
        check_row = row_idx + img_offset
        cell = sheet.cell(row=check_row, column=col_idx)
        if hasattr(sheet, '_images'):
            for image in sheet._images:
                # Check if image is in this area
                if hasattr(image, 'anchor') and hasattr(image.anchor, '_from'):
                    img_col = image.anchor._from.col
                    img_row = image.anchor._from.row
                    if img_col == col_idx - 1 and abs(img_row - (check_row-1)) < 3:
                        print(f"  ✓ Image found near row {check_row}")
                        has_image = True
                        break

    if not has_image:
        print(f"  No images detected above this CODE cell")

print(f"\n{'='*80}")
print("Analysis complete")
print(f"{'='*80}")
