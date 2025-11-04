#!/usr/bin/env python3
"""
PaddleOCR-based Excel Parser
Free, open-source, local OCR for extracting products from complex Excel layouts
"""

import sys
import json
import logging
from pathlib import Path
from typing import List, Dict, Any, Tuple
from collections import defaultdict
import re

# Excel and image processing
try:
    import openpyxl
    from PIL import Image, ImageDraw, ImageFont
    from paddleocr import PaddleOCR
    import numpy as np
except ImportError:
    print("📦 Installing required packages...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "pillow", "paddlepaddle", "paddleocr"])
    import openpyxl
    from PIL import Image, ImageDraw, ImageFont
    from paddleocr import PaddleOCR
    import numpy as np

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)


class PaddleOCRExcelParser:
    """Parse Excel files using PaddleOCR for robust local extraction"""

    def __init__(self):
        logger.info("🚀 Initializing PaddleOCR (English + Korean support)...")
        # Use 'korean' lang which supports both English and Korean characters
        self.ocr = PaddleOCR(lang='korean')
        self.output_dir = Path("data/excel_uploads/paddleocr_parsed")
        self.output_dir.mkdir(parents=True, exist_ok=True)
        logger.info("✅ PaddleOCR initialized")

    def excel_to_image(self, excel_path: Path, sheet_name: str = None,
                       start_row: int = 1, end_row: int = 100,
                       start_col: int = 1, end_col: int = 30) -> Image.Image:
        """
        Convert Excel range to high-quality image

        Args:
            excel_path: Path to Excel file
            sheet_name: Sheet name (None for active)
            start_row, end_row: Row range
            start_col, end_col: Column range

        Returns:
            PIL Image of the Excel range
        """
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb[sheet_name] if sheet_name else wb.active

        # Calculate dimensions
        col_widths = {}
        for col in range(start_col, min(end_col + 1, sheet.max_column + 1)):
            col_letter = openpyxl.utils.get_column_letter(col)
            width = sheet.column_dimensions[col_letter].width or 10
            col_widths[col] = max(width, 8) * 10  # Pixels

        row_heights = {}
        for row in range(start_row, min(end_row + 1, sheet.max_row + 1)):
            height = sheet.row_dimensions[row].height or 15
            row_heights[row] = max(height, 15) * 1.5  # Pixels

        # Create optimized image (smaller for faster OCR)
        img_width = int(sum(col_widths.values()) * 0.7)  # Reduce size by 30%
        img_height = int(sum(row_heights.values()) * 0.7)

        img = Image.new('RGB', (img_width, img_height), 'white')
        draw = ImageDraw.Draw(img)

        # Load font (optimized size for OCR)
        try:
            font = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial.ttf", 14)
            font_small = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial.ttf", 11)
        except:
            font = ImageFont.load_default()
            font_small = font

        # Draw cells with content
        y_offset = 0
        for row in range(start_row, min(end_row + 1, sheet.max_row + 1)):
            x_offset = 0
            row_height = row_heights.get(row, 20)

            for col in range(start_col, min(end_col + 1, sheet.max_column + 1)):
                cell = sheet.cell(row=row, column=col)
                col_width = col_widths.get(col, 80)

                # Draw cell border (light gray)
                draw.rectangle(
                    [x_offset, y_offset, x_offset + col_width, y_offset + row_height],
                    outline='lightgray',
                    width=1
                )

                # Draw cell value with proper font size
                if cell.value:
                    text = str(cell.value)[:100]

                    # Use larger font for important labels
                    use_font = font if any(label in text for label in ['CODE', 'SPEC', '포장', '금형']) else font_small

                    # Draw text with padding
                    draw.text(
                        (x_offset + 5, y_offset + 5),
                        text,
                        fill='black',
                        font=use_font
                    )

                x_offset += col_width

            y_offset += row_height

        wb.close()
        return img

    def extract_text_with_positions(self, image: Image.Image) -> List[Dict[str, Any]]:
        """
        Use PaddleOCR to extract text with bounding box positions

        Args:
            image: PIL Image

        Returns:
            List of text elements with positions
        """
        # Convert PIL to numpy array
        img_array = np.array(image)

        # Run OCR
        logger.info("   🔍 Running OCR analysis...")
        result = self.ocr.predict(img_array)

        if not result or not result[0]:
            logger.warning("   ⚠️ No text detected")
            return []

        # Parse OCR results
        text_elements = []
        for line in result[0]:
            bbox = line[0]  # [[x1,y1], [x2,y2], [x3,y3], [x4,y4]]
            text_info = line[1]  # (text, confidence)
            text = text_info[0]
            confidence = text_info[1]

            # Calculate center position
            x_center = sum([p[0] for p in bbox]) / 4
            y_center = sum([p[1] for p in bbox]) / 4

            text_elements.append({
                'text': text,
                'confidence': confidence,
                'bbox': bbox,
                'x': x_center,
                'y': y_center,
                'width': bbox[1][0] - bbox[0][0],
                'height': bbox[2][1] - bbox[1][1]
            })

        logger.info(f"   ✅ Detected {len(text_elements)} text elements")
        return text_elements

    def group_into_products(self, text_elements: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Group OCR text elements into product entries

        Strategy:
        1. Find all "CODE" labels
        2. For each CODE, find text to the right (product code)
        3. Find "SPEC" below CODE, and text to the right (spec value)
        4. Find pricing/production data below
        """
        products = []

        # Find CODE labels
        code_labels = [el for el in text_elements if 'CODE' in el['text'].upper() and el['confidence'] > 0.8]

        logger.info(f"   📋 Found {len(code_labels)} CODE labels")

        for code_label in code_labels:
            product = {
                'product_code': None,
                'spec': None,
                'packaging': None,
                'mold': None,
                'cost': None,
                'price': None,
                'note': None
            }

            code_y = code_label['y']
            code_x = code_label['x']

            # Find product code (to the right of CODE, same row)
            code_candidates = [
                el for el in text_elements
                if el['x'] > code_x + 50  # To the right
                and abs(el['y'] - code_y) < 30  # Same row
                and el['text'] != code_label['text']
                and el['confidence'] > 0.7
            ]

            if code_candidates:
                # Get the closest element to the right
                code_value = min(code_candidates, key=lambda el: abs(el['x'] - code_x))
                product['product_code'] = code_value['text'].strip()

            # Find SPEC label (below CODE)
            spec_labels = [
                el for el in text_elements
                if 'SPEC' in el['text'].upper()
                and el['y'] > code_y + 20  # Below CODE
                and el['y'] < code_y + 100  # Not too far
                and abs(el['x'] - code_x) < 50  # Same column
                and el['confidence'] > 0.8
            ]

            if spec_labels:
                spec_label = spec_labels[0]
                spec_y = spec_label['y']
                spec_x = spec_label['x']

                # Find spec value (to the right of SPEC)
                spec_candidates = [
                    el for el in text_elements
                    if el['x'] > spec_x + 50  # To the right
                    and abs(el['y'] - spec_y) < 30  # Same row
                    and el['text'] != spec_label['text']
                    and el['confidence'] > 0.7
                ]

                if spec_candidates:
                    spec_value = min(spec_candidates, key=lambda el: abs(el['x'] - spec_x))
                    product['spec'] = spec_value['text'].strip()

            # Find packaging (포장)
            packaging_labels = [
                el for el in text_elements
                if '포장' in el['text'] or 'packaging' in el['text'].lower()
                and el['y'] > code_y + 50
                and el['y'] < code_y + 200
                and abs(el['x'] - code_x) < 100
            ]

            if packaging_labels:
                pkg_label = packaging_labels[0]
                pkg_candidates = [
                    el for el in text_elements
                    if el['x'] > pkg_label['x'] + 30
                    and abs(el['y'] - pkg_label['y']) < 30
                    and el['confidence'] > 0.6
                ]
                if pkg_candidates:
                    pkg_value = min(pkg_candidates, key=lambda el: abs(el['x'] - pkg_label['x']))
                    product['packaging'] = pkg_value['text'].strip()

            # Only add products with at least a code or spec
            if product['product_code'] or product['spec']:
                products.append(product)

        logger.info(f"   ✅ Grouped into {len(products)} products")
        return products

    def parse_excel_file(self, excel_path: Path, rows_per_batch: int = 50) -> List[Dict[str, Any]]:
        """
        Parse entire Excel file using PaddleOCR in batches

        Args:
            excel_path: Path to Excel file
            rows_per_batch: Number of rows per batch

        Returns:
            List of all extracted products
        """
        logger.info(f"\n📊 Parsing Excel file: {excel_path.name}")

        # Detect material
        filename_upper = excel_path.name.upper()
        if 'PETG' in filename_upper:
            material = 'PETG'
        elif 'PET' in filename_upper:
            material = 'PET'
        elif 'PE' in filename_upper:
            material = 'PE'
        elif 'PP' in filename_upper:
            material = 'PP'
        else:
            material = 'Other'

        # Get sheet dimensions
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb.active
        total_rows = sheet.max_row
        wb.close()

        logger.info(f"   Material: {material}")
        logger.info(f"   Total rows: {total_rows}")

        all_products = []

        # Process first batch as test
        start_row = 1
        end_row = min(rows_per_batch, total_rows)

        logger.info(f"\n   Processing rows {start_row}-{end_row}...")

        # Convert to image
        image = self.excel_to_image(excel_path, start_row=start_row, end_row=end_row)

        # Save image
        debug_image_path = self.output_dir / f"{excel_path.stem}_rows_{start_row}_{end_row}.png"
        image.save(debug_image_path)
        logger.info(f"   📷 Saved image: {debug_image_path.name}")

        # Extract text with OCR
        text_elements = self.extract_text_with_positions(image)

        # Group into products
        if text_elements:
            products = self.group_into_products(text_elements)

            # Add material to each product
            for product in products:
                product['material'] = material

            all_products.extend(products)

        logger.info(f"\n✅ Total products extracted: {len(all_products)}")

        # Save results
        output_file = self.output_dir / f"{excel_path.stem}_products.json"
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(all_products, f, ensure_ascii=False, indent=2)

        logger.info(f"💾 Saved to: {output_file}")

        # Save OCR debug info
        debug_file = self.output_dir / f"{excel_path.stem}_ocr_elements.json"
        with open(debug_file, 'w', encoding='utf-8') as f:
            json.dump(text_elements, f, ensure_ascii=False, indent=2)

        logger.info(f"🔍 Debug info: {debug_file}")

        return all_products


def main():
    """Main execution - test with small batch"""

    parser = PaddleOCRExcelParser()

    # Test with PE file
    excel_file = Path("data/excel_uploads/raw/제품 리스트_1.PE.xlsx")

    if not excel_file.exists():
        print(f"❌ Excel file not found: {excel_file}")
        return

    print("\n🚀 Starting PaddleOCR-based Excel parsing...")
    print("📝 This uses FREE local OCR (no API costs)\n")

    # Parse first 20 rows as test (optimized)
    products = parser.parse_excel_file(excel_file, rows_per_batch=20)

    print(f"\n{'='*80}")
    print(f"📊 Parsing Summary:")
    print(f"   File: {excel_file.name}")
    print(f"   Total products: {len(products)}")
    print(f"   Output: {parser.output_dir}")
    print(f"{'='*80}")

    # Show samples
    if products:
        print(f"\n📝 Sample products (first 10):")
        for i, product in enumerate(products[:10], 1):
            print(f"\n{i}. Code: {product.get('product_code', 'N/A')}")
            print(f"   Spec: {product.get('spec', 'N/A')}")
            print(f"   Material: {product.get('material', 'N/A')}")
            print(f"   Packaging: {product.get('packaging', 'N/A')}")


if __name__ == "__main__":
    main()
