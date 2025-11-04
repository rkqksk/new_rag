#!/usr/bin/env python3
"""
Excel Document Intelligence Parser
Hybrid approach: Structure-based + OCR + Image extraction
"""

import sys
import json
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from collections import defaultdict
import io

# UTF-8 강제 설정
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

try:
    import openpyxl
    from PIL import Image
    from paddleocr import PaddleOCR
    import numpy as np
except ImportError as e:
    print(f"❌ Missing dependencies: {e}")
    print("Run: pip install openpyxl pillow paddlepaddle paddleocr")
    sys.exit(1)

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)


class ExcelDocumentParser:
    """
    Intelligent Excel parser with automatic strategy selection

    Features:
    - Pattern detection and structure analysis
    - Hybrid parsing (openpyxl + OCR)
    - Image and chart extraction
    - UTF-8 clean output
    """

    def __init__(self, use_ocr: bool = False):
        """
        Args:
            use_ocr: Enable OCR for complex layouts (slower but more robust)
        """
        self.use_ocr = use_ocr
        if use_ocr:
            logger.info("🚀 Initializing PaddleOCR (Korean + English)...")
            self.ocr = PaddleOCR(lang='korean')
            logger.info("✅ OCR ready")
        else:
            self.ocr = None

        self.output_dir = Path("data/excel_uploads/parsed")
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def detect_material(self, filename: str) -> str:
        """Detect material type from filename"""
        filename_upper = filename.upper()
        if 'PETG' in filename_upper:
            return 'PETG'
        elif 'PET' in filename_upper and 'PETG' not in filename_upper:
            return 'PET'
        elif 'PE' in filename_upper:
            return 'PE'
        elif 'PP' in filename_upper or 'PS' in filename_upper or 'ABS' in filename_upper:
            return 'PP'
        else:
            return 'Other'

    def detect_pattern(self, sheet) -> Optional[Dict[str, Any]]:
        """
        Detect repeating patterns in Excel sheet

        Returns:
            Pattern info if found, None otherwise
        """
        logger.info("   🔍 Detecting sheet structure pattern...")

        # Find all CODE cells (check entire sheet)
        code_cells = []
        max_check_rows = sheet.max_row

        for row in sheet.iter_rows(min_row=1, max_row=max_check_rows):
            for cell in row:
                if cell.value == "CODE":
                    code_cells.append(cell)

        logger.info(f"   Found {len(code_cells)} CODE cells in first {max_check_rows} rows")

        if len(code_cells) < 3:
            logger.warning(f"   ⚠️ Not enough CODE cells to detect pattern ({len(code_cells)} found, need 3+)")
            return None

        # Group CODE cells by column (handle horizontal grid layout)
        cells_by_column = defaultdict(list)
        for cell in code_cells:
            cells_by_column[cell.column].append(cell)

        # Find the column with the most CODE cells (main column)
        main_column = max(cells_by_column.keys(), key=lambda col: len(cells_by_column[col]))
        main_column_cells = sorted(cells_by_column[main_column], key=lambda c: c.row)

        logger.info(f"   Column analysis: {len(cells_by_column)} columns, main column={main_column} with {len(main_column_cells)} CODE cells")

        if len(main_column_cells) < 3:
            logger.warning(f"   ⚠️ Not enough CODE cells in main column ({len(main_column_cells)} found, need 3+)")
            return None

        # Calculate distances between CODE cells in the main column only
        distances = []
        for i in range(len(main_column_cells) - 1):
            dist = main_column_cells[i + 1].row - main_column_cells[i].row
            distances.append(dist)

        # Check if pattern is consistent
        if distances:
            # Use median instead of average for better robustness
            sorted_distances = sorted(distances)
            median_dist = sorted_distances[len(sorted_distances) // 2]

            # Count how many distances match the median (with ±1 tolerance)
            matching = sum(1 for d in distances if abs(d - median_dist) <= 1)
            consistency_ratio = matching / len(distances)

            logger.info(f"   Distance analysis: median={median_dist}, consistency={consistency_ratio:.1%}")

            # If 70%+ of distances match the pattern, consider it valid
            if consistency_ratio >= 0.7:
                pattern = {
                    'type': 'repeating_blocks',
                    'block_size': int(median_dist),
                    'anchor_label': 'CODE',
                    'anchor_count': len(main_column_cells),
                    'first_anchor_row': main_column_cells[0].row,
                    'anchor_column': main_column,
                    'consistency': consistency_ratio,
                    'total_code_cells': len(code_cells),
                    'columns_with_code': len(cells_by_column)
                }
                logger.info(f"   ✅ Pattern detected: {pattern['block_size']}-row blocks, {pattern['anchor_count']} products in column {main_column}")
                return pattern
            else:
                logger.info(f"   ℹ️ Pattern too inconsistent ({consistency_ratio:.1%} match)")

        logger.info("   ℹ️ No consistent pattern detected")
        return None

    def extract_images(self, sheet, output_subdir: str) -> List[Dict[str, Any]]:
        """Extract all embedded images from sheet"""
        images_data = []

        if not hasattr(sheet, '_images') or not sheet._images:
            logger.info("   ℹ️ No images found in sheet")
            return images_data

        logger.info(f"   📷 Extracting {len(sheet._images)} images...")

        image_dir = self.output_dir / output_subdir / "images"
        image_dir.mkdir(parents=True, exist_ok=True)

        for idx, img in enumerate(sheet._images, 1):
            try:
                # Get image position
                if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                    img_row = img.anchor._from.row + 1  # 0-based to 1-based
                    img_col = img.anchor._from.col + 1
                else:
                    img_row, img_col = 0, 0

                # Save image
                image_path = image_dir / f"image_{idx:03d}_r{img_row}c{img_col}.png"

                # Convert to PIL Image and save
                if hasattr(img, '_data'):
                    pil_img = Image.open(io.BytesIO(img._data()))
                    pil_img.save(image_path)

                    images_data.append({
                        'index': idx,
                        'row': img_row,
                        'column': img_col,
                        'path': str(image_path),
                        'size': {'width': pil_img.width, 'height': pil_img.height}
                    })

            except Exception as e:
                logger.warning(f"   ⚠️ Failed to extract image {idx}: {e}")

        logger.info(f"   ✅ Extracted {len(images_data)} images to {image_dir}")
        return images_data

    def parse_block_pattern(self, sheet, pattern: Dict[str, Any], material: str) -> List[Dict[str, Any]]:
        """
        Parse Excel using detected block pattern

        For 11-row pattern:
        - Row 0: CODE label
        - Row 1: Product code
        - Row 2: SPEC label
        - Row 3: Spec value
        - Row 4: Dimensions
        - Row 5: 포장 (Packaging)
        - Row 6: 금형 (Mold)
        - Row 7: 원가 (Cost)
        - Row 8: 판매 (Price)
        - Row 9: 생산량 (Production)
        - Row 10: 비고 (Note)
        """
        products = []
        anchor_col = pattern['anchor_column']
        block_size = pattern['block_size']
        first_row = pattern['first_anchor_row']

        logger.info(f"   📋 Parsing {pattern['anchor_count']} product blocks...")

        for block_idx in range(pattern['anchor_count']):
            base_row = first_row + (block_idx * block_size)

            product = {
                'material': material,
                'product_code': None,
                'spec': None,
                'dimensions': None,
                'packaging': None,
                'mold': None,
                'cost': None,
                'price': None,
                'production': None,
                'note': None,
                'source_row': base_row
            }

            try:
                # Extract data with relative offsets from CODE cell
                # CODE is at base_row, data is typically in next column
                data_col = anchor_col + 1

                # Product code (row after CODE)
                code_val = sheet.cell(row=base_row + 1, column=data_col).value
                if code_val:
                    product['product_code'] = str(code_val).strip()

                # SPEC value (2-3 rows after CODE)
                spec_val = sheet.cell(row=base_row + 2, column=data_col).value
                if not spec_val:
                    spec_val = sheet.cell(row=base_row + 3, column=data_col).value
                if spec_val:
                    product['spec'] = str(spec_val).strip()

                # Dimensions (3-4 rows after CODE)
                dim_val = sheet.cell(row=base_row + 4, column=data_col).value
                if dim_val:
                    product['dimensions'] = str(dim_val).strip()

                # Packaging (포장 - typically 5 rows after)
                pkg_val = sheet.cell(row=base_row + 5, column=data_col).value
                if pkg_val:
                    product['packaging'] = str(pkg_val).strip()

                # Mold (금형 - 6 rows after)
                mold_val = sheet.cell(row=base_row + 6, column=data_col).value
                if mold_val:
                    product['mold'] = str(mold_val).strip()

                # Cost (원가 - 7 rows after)
                cost_val = sheet.cell(row=base_row + 7, column=data_col).value
                if cost_val and isinstance(cost_val, (int, float)):
                    product['cost'] = float(cost_val)

                # Price (판매 - 8 rows after)
                price_val = sheet.cell(row=base_row + 8, column=data_col).value
                if price_val and isinstance(price_val, (int, float)):
                    product['price'] = float(price_val)

                # Production (생산량 - 9 rows after)
                prod_val = sheet.cell(row=base_row + 9, column=data_col).value
                if prod_val:
                    product['production'] = str(prod_val).strip()

                # Note (비고 - 10 rows after)
                note_val = sheet.cell(row=base_row + 10, column=data_col).value
                if note_val:
                    product['note'] = str(note_val).strip()

                # Only add if we have at least product code or spec
                if product['product_code'] or product['spec']:
                    products.append(product)

            except Exception as e:
                logger.warning(f"   ⚠️ Error parsing block {block_idx} at row {base_row}: {e}")

        logger.info(f"   ✅ Parsed {len(products)} products")
        return products

    def parse_excel_file(self, excel_path: Path) -> Dict[str, Any]:
        """
        Main parsing method - automatically selects best strategy

        Returns:
            Parsing results with products, images, and metadata
        """
        logger.info(f"\n📊 Parsing: {excel_path.name}")

        # Detect material
        material = self.detect_material(excel_path.name)
        logger.info(f"   Material: {material}")

        # Load workbook
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb.active
        logger.info(f"   Sheet: {sheet.title}, {sheet.max_row} rows x {sheet.max_column} cols")

        # Create output subdirectory
        output_subdir = excel_path.stem

        # Extract images first
        images = self.extract_images(sheet, output_subdir)

        # Detect pattern
        pattern = self.detect_pattern(sheet)

        products = []
        parsing_method = "unknown"

        if pattern and pattern['type'] == 'repeating_blocks':
            # Use pattern-based parsing (fastest and most accurate)
            parsing_method = "pattern-based"
            products = self.parse_block_pattern(sheet, pattern, material)

        elif self.use_ocr:
            # Fall back to OCR if no pattern detected
            parsing_method = "ocr-based"
            logger.info("   🔍 Using OCR fallback...")
            # TODO: Implement OCR fallback
            products = []

        else:
            # Basic row-by-row parsing
            parsing_method = "row-based"
            logger.warning("   ⚠️ No pattern detected, using basic parsing")
            products = []

        wb.close()

        result = {
            'file': excel_path.name,
            'material': material,
            'parsing_method': parsing_method,
            'pattern': pattern,
            'products': products,
            'images': images,
            'stats': {
                'total_products': len(products),
                'total_images': len(images),
                'sheet_rows': sheet.max_row,
                'sheet_cols': sheet.max_column
            }
        }

        # Save results
        result_file = self.output_dir / f"{excel_path.stem}_parsed.json"
        with open(result_file, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)

        logger.info(f"   💾 Saved to: {result_file}")
        logger.info(f"   ✅ Complete: {len(products)} products, {len(images)} images")

        return result


def main():
    """Test the document parser"""

    parser = ExcelDocumentParser(use_ocr=False)

    # Test with PE file
    excel_file = Path("data/excel_uploads/raw/제품 리스트_1.PE.xlsx")

    if not excel_file.exists():
        print(f"❌ File not found: {excel_file}")
        return

    print("="* 80)
    print("Excel Document Intelligence Parser")
    print("="* 80)

    result = parser.parse_excel_file(excel_file)

    print(f"\n{'='*80}")
    print("Parsing Summary:")
    print(f"{'='*80}")
    print(f"  File: {result['file']}")
    print(f"  Material: {result['material']}")
    print(f"  Method: {result['parsing_method']}")
    print(f"  Products: {result['stats']['total_products']}")
    print(f"  Images: {result['stats']['total_images']}")
    print(f"{'='*80}")

    # Show first 5 products
    if result['products']:
        print(f"\n📝 Sample products (first 5):")
        for i, product in enumerate(result['products'][:5], 1):
            print(f"\n{i}. {product.get('product_code', 'N/A')}")
            print(f"   Spec: {product.get('spec', 'N/A')}")
            print(f"   Material: {product.get('material', 'N/A')}")
            print(f"   Packaging: {product.get('packaging', 'N/A')}")
            if product.get('cost'):
                print(f"   Cost: {product['cost']:.2f}")
            if product.get('price'):
                print(f"   Price: {product['price']:.2f}")


if __name__ == "__main__":
    main()
