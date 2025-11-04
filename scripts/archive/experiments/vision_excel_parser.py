#!/usr/bin/env python3
"""
Vision-Based Excel Parser
Uses OCR and Vision AI to extract product data from complex Excel layouts
"""

import sys
import json
import base64
import logging
from pathlib import Path
from typing import List, Dict, Any
from io import BytesIO

# Excel and image processing
try:
    import openpyxl
    from PIL import Image, ImageDraw, ImageFont
    import anthropic
except ImportError:
    print("Installing required packages...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "pillow", "anthropic"])
    import openpyxl
    from PIL import Image, ImageDraw, ImageFont
    import anthropic

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)


class VisionExcelParser:
    """Parse Excel files using vision AI for robust extraction"""

    def __init__(self, api_key: str = None):
        self.api_key = api_key or self._get_api_key()
        self.client = anthropic.Anthropic(api_key=self.api_key)
        self.output_dir = Path("data/excel_uploads/vision_parsed")
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _get_api_key(self) -> str:
        """Get Anthropic API key from environment"""
        import os
        api_key = os.getenv("ANTHROPIC_API_KEY")
        if not api_key:
            raise ValueError("ANTHROPIC_API_KEY environment variable not set")
        return api_key

    def excel_to_image(self, excel_path: Path, sheet_name: str = None,
                       start_row: int = 1, end_row: int = 50) -> Image.Image:
        """
        Convert Excel range to image for vision processing

        Args:
            excel_path: Path to Excel file
            sheet_name: Sheet name (None for active)
            start_row: Starting row to capture
            end_row: Ending row to capture

        Returns:
            PIL Image of the Excel range
        """
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb[sheet_name] if sheet_name else wb.active

        # Calculate image dimensions
        col_widths = {}
        for col in range(1, sheet.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            col_widths[col] = max(
                sheet.column_dimensions[col_letter].width or 10,
                10
            ) * 8  # Convert to pixels

        row_heights = {}
        for row in range(start_row, min(end_row + 1, sheet.max_row + 1)):
            row_heights[row] = max(
                sheet.row_dimensions[row].height or 15,
                15
            ) * 1.3  # Convert to pixels

        # Create image
        img_width = sum(col_widths.values())
        img_height = sum(row_heights.values())

        img = Image.new('RGB', (int(img_width), int(img_height)), 'white')
        draw = ImageDraw.Draw(img)

        # Try to load a font
        try:
            font = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial.ttf", 12)
        except:
            font = ImageFont.load_default()

        # Draw cells
        y_offset = 0
        for row in range(start_row, min(end_row + 1, sheet.max_row + 1)):
            x_offset = 0
            row_height = row_heights.get(row, 20)

            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                col_width = col_widths.get(col, 80)

                # Draw cell border
                draw.rectangle(
                    [x_offset, y_offset, x_offset + col_width, y_offset + row_height],
                    outline='black',
                    width=1
                )

                # Draw cell value
                if cell.value:
                    text = str(cell.value)[:50]  # Truncate long text
                    draw.text(
                        (x_offset + 5, y_offset + 5),
                        text,
                        fill='black',
                        font=font
                    )

                x_offset += col_width

            y_offset += row_height

        wb.close()
        return img

    def image_to_base64(self, image: Image.Image) -> str:
        """Convert PIL Image to base64 string"""
        buffered = BytesIO()
        image.save(buffered, format="PNG")
        return base64.b64encode(buffered.getvalue()).decode('utf-8')

    def extract_products_from_image(self, image: Image.Image, material: str) -> List[Dict[str, Any]]:
        """
        Use Claude Vision API to extract product data from Excel image

        Args:
            image: PIL Image of Excel sheet
            material: Material type (PE, PET, PETG, PP, Other)

        Returns:
            List of extracted products
        """
        logger.info(f"🔍 Analyzing image with Claude Vision API...")

        # Convert image to base64
        image_b64 = self.image_to_base64(image)

        # Create prompt for vision extraction
        prompt = f"""Analyze this Excel product catalog image and extract ALL products with their details.

This is a {material} product list. Each product block contains:
- CODE: Product code (format: XX000-X000 or internal code)
- SPEC: Specification (material / capacity)
- Images: Product photos
- 포장 (Packaging): Packaging details
- 금형 (Mold): Mold information
- 원가 (Cost): Cost data
- 판매 (Price): Selling price
- Other details

Extract each product as a JSON object with fields:
- product_code: The CODE value (if empty, mark as null)
- spec: The SPEC value
- material: "{material}"
- packaging: 포장 value
- mold: 금형 value
- cost: 원가 value (number or null)
- price: 판매 value (number or null)
- production: 생산량 value if present
- note: Any notes/remarks

Return ONLY a JSON array of products, no other text:
[
  {{"product_code": "BE010-GU01", "spec": "PE / 10ml", "material": "{material}", ...}},
  {{"product_code": "BE020-G003", "spec": "PE / 20ml", "material": "{material}", ...}},
  ...
]

Important:
- Extract ALL products visible in the image
- If a field is empty/missing, use null
- Product codes can be in various formats (BE010-GU01, OPE 40ml _20, etc.)
- Some CODE cells may be empty - still extract those products with code as null
"""

        try:
            message = self.client.messages.create(
                model="claude-3-5-sonnet-20241022",
                max_tokens=4096,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "image",
                                "source": {
                                    "type": "base64",
                                    "media_type": "image/png",
                                    "data": image_b64,
                                },
                            },
                            {
                                "type": "text",
                                "text": prompt
                            }
                        ],
                    }
                ],
            )

            # Extract JSON from response
            response_text = message.content[0].text

            # Try to find JSON array in response
            import re
            json_match = re.search(r'\[.*\]', response_text, re.DOTALL)
            if json_match:
                products_json = json_match.group()
                products = json.loads(products_json)
                logger.info(f"✅ Extracted {len(products)} products")
                return products
            else:
                logger.error(f"No JSON array found in response")
                return []

        except Exception as e:
            logger.error(f"Error in vision extraction: {e}")
            return []

    def parse_excel_file(self, excel_path: Path, rows_per_batch: int = 40) -> List[Dict[str, Any]]:
        """
        Parse entire Excel file using vision API in batches

        Args:
            excel_path: Path to Excel file
            rows_per_batch: Number of rows to process per vision API call

        Returns:
            List of all extracted products
        """
        logger.info(f"\n📊 Parsing Excel file: {excel_path.name}")

        # Detect material from filename
        filename_upper = excel_path.name.upper()
        if 'PETG' in filename_upper:
            material = 'PETG'
        elif 'PET' in filename_upper:
            material = 'PET'
        elif 'PE' in filename_upper:
            material = 'PE'
        elif 'PP' in filename_upper or 'PS' in filename_upper or 'ABS' in filename_upper:
            material = 'PP'
        else:
            material = 'Other'

        # Get sheet dimensions
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb.active
        total_rows = sheet.max_row
        wb.close()

        logger.info(f"   Material: {material}")
        logger.info(f"   Total rows: {total_rows}")
        logger.info(f"   Processing in batches of {rows_per_batch} rows...")

        all_products = []

        # Process in batches
        for start_row in range(1, total_rows + 1, rows_per_batch):
            end_row = min(start_row + rows_per_batch - 1, total_rows)

            logger.info(f"\n   Batch: rows {start_row}-{end_row}")

            # Convert batch to image
            image = self.excel_to_image(excel_path, start_row=start_row, end_row=end_row)

            # Save image for debugging
            debug_image_path = self.output_dir / f"{excel_path.stem}_rows_{start_row}_{end_row}.png"
            image.save(debug_image_path)
            logger.info(f"   📷 Saved image: {debug_image_path.name}")

            # Extract products from image
            products = self.extract_products_from_image(image, material)

            if products:
                all_products.extend(products)
                logger.info(f"   ✅ Extracted {len(products)} products from this batch")
            else:
                logger.warning(f"   ⚠️ No products extracted from this batch")

        logger.info(f"\n✅ Total products extracted from {excel_path.name}: {len(all_products)}")

        # Save results
        output_file = self.output_dir / f"{excel_path.stem}_products.json"
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(all_products, f, ensure_ascii=False, indent=2)

        logger.info(f"💾 Saved to: {output_file}")

        return all_products


def main():
    """Main execution - test with one file first"""

    # Check for API key
    import os
    if not os.getenv("ANTHROPIC_API_KEY"):
        print("❌ Error: ANTHROPIC_API_KEY environment variable not set")
        print("Set it with: export ANTHROPIC_API_KEY='your-key-here'")
        return

    parser = VisionExcelParser()

    # Test with PE file first
    excel_file = Path("data/excel_uploads/raw/제품 리스트_1.PE.xlsx")

    if not excel_file.exists():
        print(f"❌ Excel file not found: {excel_file}")
        return

    print("🚀 Starting vision-based Excel parsing...")
    print("📝 This will use Claude Vision API to extract products")
    print()

    # Parse first 100 rows as test
    products = parser.parse_excel_file(excel_file, rows_per_batch=50)

    print(f"\n{'='*80}")
    print(f"📊 Parsing Summary:")
    print(f"   File: {excel_file.name}")
    print(f"   Total products: {len(products)}")
    print(f"   Output: {parser.output_dir}")
    print(f"{'='*80}")

    # Show sample
    if products:
        print(f"\n📝 Sample products (first 5):")
        for i, product in enumerate(products[:5], 1):
            print(f"\n{i}. {product.get('product_code', 'N/A')}")
            print(f"   Spec: {product.get('spec', 'N/A')}")
            print(f"   Material: {product.get('material', 'N/A')}")


if __name__ == "__main__":
    main()
