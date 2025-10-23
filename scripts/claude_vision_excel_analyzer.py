#!/usr/bin/env python3
"""
Claude Vision-Based Excel Document Analyzer
Learn document patterns using Claude's vision capabilities
"""

import sys
import json
import base64
import logging
from pathlib import Path
from typing import List, Dict, Any
from io import BytesIO
import io

# UTF-8 강제 설정
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

try:
    import openpyxl
    from PIL import Image, ImageDraw, ImageFont
    import anthropic
    import os
except ImportError as e:
    print(f"❌ Missing dependencies: {e}")
    print("Run: pip install openpyxl pillow anthropic")
    sys.exit(1)

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)


class ClaudeVisionExcelAnalyzer:
    """
    Use Claude Vision API to understand and learn Excel document patterns
    """

    def __init__(self):
        api_key = os.getenv("ANTHROPIC_API_KEY")
        if not api_key:
            raise ValueError("ANTHROPIC_API_KEY environment variable not set")

        self.client = anthropic.Anthropic(api_key=api_key)
        self.output_dir = Path("data/excel_uploads/vision_analysis")
        self.output_dir.mkdir(parents=True, exist_ok=True)

        logger.info("✅ Claude Vision Analyzer initialized")

    def excel_to_image_korean_support(
        self,
        excel_path: Path,
        sheet_name: str = None,
        start_row: int = 1,
        end_row: int = 30,
        start_col: int = 1,
        end_col: int = None  # None = auto-detect
    ) -> Image.Image:
        """
        Convert Excel to image with Korean font support

        If end_col is None, automatically detects the maximum column from sheet
        """
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb[sheet_name] if sheet_name else wb.active

        # Auto-detect end_col if not provided
        if end_col is None:
            end_col = sheet.max_column
            logger.info(f"   Auto-detected max column: {end_col} (column {openpyxl.utils.get_column_letter(end_col)})")

        # Calculate dimensions
        col_widths = {}
        for col in range(start_col, min(end_col + 1, sheet.max_column + 1)):
            col_letter = openpyxl.utils.get_column_letter(col)
            width = sheet.column_dimensions[col_letter].width or 10
            col_widths[col] = max(width, 8) * 10

        row_heights = {}
        for row in range(start_row, min(end_row + 1, sheet.max_row + 1)):
            height = sheet.row_dimensions[row].height or 15
            row_heights[row] = max(height, 15) * 1.5

        img_width = int(sum(col_widths.values()))
        img_height = int(sum(row_heights.values()))

        img = Image.new('RGB', (img_width, img_height), 'white')
        draw = ImageDraw.Draw(img)

        # Korean font support
        korean_fonts = [
            "/System/Library/Fonts/AppleSDGothicNeo.ttc",
            "/System/Library/Fonts/AppleGothic.ttf",
            "/System/Library/Fonts/Supplemental/Arial Unicode.ttf"
        ]

        font = None
        font_small = None

        for font_path in korean_fonts:
            try:
                font = ImageFont.truetype(font_path, 16)
                font_small = ImageFont.truetype(font_path, 13)
                logger.info(f"   Using Korean font: {Path(font_path).name}")
                break
            except:
                continue

        if not font:
            logger.warning("   Korean font not found, using default")
            font = ImageFont.load_default()
            font_small = font

        # Draw cells
        y_offset = 0
        for row in range(start_row, min(end_row + 1, sheet.max_row + 1)):
            x_offset = 0
            row_height = row_heights.get(row, 20)

            for col in range(start_col, min(end_col + 1, sheet.max_column + 1)):
                cell = sheet.cell(row=row, column=col)
                col_width = col_widths.get(col, 80)

                # Cell border
                draw.rectangle(
                    [x_offset, y_offset, x_offset + col_width, y_offset + row_height],
                    outline='gray',
                    width=1
                )

                # Cell value with Korean support
                if cell.value:
                    text = str(cell.value)[:100]

                    # Use larger font for labels
                    use_font = font if any(label in text for label in ['CODE', 'SPEC', '포장', '금형', '원가', '판매', '생산량', '비고']) else font_small

                    # Draw text
                    draw.text(
                        (x_offset + 5, y_offset + 5),
                        text,
                        fill='black',
                        font=use_font
                    )

                x_offset += col_width

            y_offset += row_height

        wb.close()
        return img

    def image_to_base64(self, image: Image.Image) -> str:
        """Convert PIL Image to base64"""
        buffered = BytesIO()
        image.save(buffered, format="PNG")
        return base64.b64encode(buffered.getvalue()).decode('utf-8')

    def analyze_document_structure(self, image: Image.Image, material: str) -> Dict[str, Any]:
        """
        Use Claude Vision to understand document structure and extract data

        This is the KEY method - Claude learns the pattern
        """
        logger.info("🔍 Analyzing document with Claude Vision API...")

        image_b64 = self.image_to_base64(image)

        prompt = f"""Analyze this Excel product catalog image in detail. This is a {material} product list.

**Your task**: Extract ALL product information and learn the document structure pattern.

**Expected structure analysis:**
1. Identify all label fields (Korean and English) - like CODE, SPEC, 포장, 금형, etc.
2. Understand the layout pattern - how products are organized (vertical blocks, horizontal grid, etc.)
3. Extract every piece of data with its correct label

**Output format** - JSON array with this exact structure:
```json
[
  {{
    "product_code": "exact code from CODE field",
    "spec": "material / capacity",
    "dimensions": "size information if present",
    "packaging_label": "포장 or packaging",
    "packaging_value": "packaging data",
    "mold_label": "금형 or mold",
    "mold_value": "mold data",
    "cost_label": "원가 or cost",
    "cost_value": cost as number or null,
    "price_label": "판매 or price",
    "price_value": price as number or null,
    "production_label": "생산량 or production",
    "production_value": "production data",
    "note_label": "비고 or note",
    "note_value": "note/remark data",
    "material": "{material}"
  }}
]
```

**Critical requirements:**
- Extract EVERY product visible in the image
- Preserve exact Korean labels (포장, 금형, 원가, 판매, 생산량, 비고)
- If a field is empty, use null (not empty string)
- Numbers should be actual numbers, not strings
- Be precise with product codes (e.g., BE040-R001)

**Pattern learning notes:**
After extraction, describe:
1. Layout pattern (how many rows per product?)
2. Label positions (left column labels vs data columns?)
3. Repetition structure (vertical stack vs horizontal grid?)
4. Any irregularities or special cases

Return ONLY valid JSON, no other text."""

        try:
            message = self.client.messages.create(
                model="claude-3-5-sonnet-20241022",
                max_tokens=8192,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "image",
                                "source": {
                                    "type": "base64",
                                    "media_type": "image/png",
                                    "data": image_b64,
                                },
                            },
                            {
                                "type": "text",
                                "text": prompt
                            }
                        ],
                    }
                ],
            )

            response_text = message.content[0].text
            logger.info(f"✅ Received response from Claude Vision API")

            # Extract JSON
            import re
            json_match = re.search(r'```json\s*(.*?)\s*```', response_text, re.DOTALL)
            if not json_match:
                json_match = re.search(r'\[.*\]', response_text, re.DOTALL)

            if json_match:
                products_json = json_match.group(1) if json_match.lastindex else json_match.group()
                products = json.loads(products_json)

                # Extract pattern notes if present
                pattern_notes = response_text.split('```')[-1] if '```' in response_text else ""

                result = {
                    'products': products,
                    'pattern_analysis': pattern_notes.strip(),
                    'raw_response': response_text
                }

                logger.info(f"✅ Extracted {len(products)} products")
                return result
            else:
                logger.error("❌ No JSON found in response")
                return {
                    'products': [],
                    'pattern_analysis': '',
                    'raw_response': response_text
                }

        except Exception as e:
            logger.error(f"❌ Vision API error: {e}")
            return {
                'products': [],
                'pattern_analysis': '',
                'error': str(e)
            }

    def analyze_excel_file(
        self,
        excel_path: Path,
        rows_per_batch: int = 30,
        num_batches: int = 3
    ) -> Dict[str, Any]:
        """
        Analyze Excel file using Claude Vision in batches
        """
        logger.info(f"\n📊 Analyzing Excel file: {excel_path.name}")

        # Detect material
        filename_upper = excel_path.name.upper()
        if 'PETG' in filename_upper:
            material = 'PETG'
        elif 'PET' in filename_upper:
            material = 'PET'
        elif 'PE' in filename_upper:
            material = 'PE'
        elif 'PP' in filename_upper:
            material = 'PP'
        else:
            material = 'Other'

        logger.info(f"   Material: {material}")

        all_products = []
        all_patterns = []

        # Process multiple batches to learn patterns
        for batch_idx in range(num_batches):
            start_row = 1 + (batch_idx * rows_per_batch)
            end_row = start_row + rows_per_batch - 1

            logger.info(f"\n   📸 Batch {batch_idx + 1}/{num_batches}: rows {start_row}-{end_row}")

            # Generate image
            image = self.excel_to_image_korean_support(
                excel_path,
                start_row=start_row,
                end_row=end_row
            )

            # Save image
            image_path = self.output_dir / f"{excel_path.stem}_batch{batch_idx + 1}.png"
            image.save(image_path)
            logger.info(f"   💾 Saved: {image_path.name}")

            # Analyze with Vision API
            result = self.analyze_document_structure(image, material)

            if result['products']:
                all_products.extend(result['products'])
                logger.info(f"   ✅ Batch {batch_idx + 1}: {len(result['products'])} products")

            if result.get('pattern_analysis'):
                all_patterns.append({
                    'batch': batch_idx + 1,
                    'rows': f"{start_row}-{end_row}",
                    'pattern': result['pattern_analysis']
                })

        # Save comprehensive results
        output_file = self.output_dir / f"{excel_path.stem}_vision_analysis.json"
        comprehensive_result = {
            'file': excel_path.name,
            'material': material,
            'total_products': len(all_products),
            'products': all_products,
            'pattern_learning': all_patterns
        }

        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(comprehensive_result, f, ensure_ascii=False, indent=2)

        logger.info(f"\n💾 Saved comprehensive analysis: {output_file}")
        logger.info(f"✅ Total products extracted: {len(all_products)}")

        return comprehensive_result


def main():
    """Test Claude Vision analyzer"""

    analyzer = ClaudeVisionExcelAnalyzer()

    excel_file = Path("data/excel_uploads/raw/제품 리스트_1.PE.xlsx")

    if not excel_file.exists():
        print(f"❌ File not found: {excel_file}")
        return

    print("=" * 80)
    print("Claude Vision Excel Document Analyzer")
    print("Learning document patterns with AI")
    print("=" * 80)

    # Analyze with Vision API (first 90 rows, 3 batches)
    result = analyzer.analyze_excel_file(excel_file, rows_per_batch=30, num_batches=3)

    print(f"\n{'='*80}")
    print("Analysis Complete!")
    print(f"{'='*80}")
    print(f"  Total products: {result['total_products']}")
    print(f"  Pattern batches: {len(result['pattern_learning'])}")
    print(f"{'='*80}")

    # Show sample products
    if result['products']:
        print(f"\n📝 Sample products (first 5):")
        for i, product in enumerate(result['products'][:5], 1):
            print(f"\n{i}. {product.get('product_code', 'N/A')}")
            print(f"   Spec: {product.get('spec', 'N/A')}")
            print(f"   포장: {product.get('packaging_value', 'N/A')}")
            print(f"   원가: {product.get('cost_value', 'N/A')}")
            print(f"   판매: {product.get('price_value', 'N/A')}")

    # Show pattern learning
    if result['pattern_learning']:
        print(f"\n🧠 Pattern Learning Summary:")
        for pattern_info in result['pattern_learning']:
            print(f"\nBatch {pattern_info['batch']} ({pattern_info['rows']}):")
            print(f"  {pattern_info['pattern'][:200]}...")


if __name__ == "__main__":
    main()
