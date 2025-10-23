#!/usr/bin/env python3
"""
Generate Master Comparison Excel with Visual Indicators
Creates comprehensive Excel workbook with color-coded cells for missing data
"""

import sys
import json
import csv
import logging
from pathlib import Path
from typing import Dict, List, Any
from collections import defaultdict

# Excel formatting
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl not installed. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from app.services.excel_parser_service import ExcelParserService

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)


class MasterComparisonGenerator:
    """Generate comprehensive Excel comparison with visual indicators"""

    def __init__(self):
        self.crawled_base = Path("data/crawled_products_final")
        self.excel_base = Path("data/excel_uploads")

        # Color schemes
        self.RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self.YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.GREEN_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        self.ORANGE_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        self.LIGHT_GRAY_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        self.WHITE_FONT = Font(color="FFFFFF", bold=True)
        self.BOLD_FONT = Font(bold=True)
        self.HEADER_FONT = Font(bold=True, size=12)

        self.CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

        self.THIN_BORDER = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def load_all_data(self):
        """Load all product data from both sources"""

        logger.info("📊 Loading all product data...")

        # Load crawled data
        crawled_products = {}
        for json_file in self.crawled_base.rglob("*.json"):
            if json_file.name.startswith("idx_"):
                try:
                    with open(json_file, 'r', encoding='utf-8') as f:
                        data = json.load(f)

                    specs = data.get("specifications", {})
                    code = specs.get("제품 코드", "N/A")

                    if code and code != "N/A":
                        crawled_products[code] = {
                            "source": "crawled",
                            "product_name": data.get("product_name", ""),
                            "product_code": code,
                            "material": specs.get("재질(원료)", ""),
                            "spec": specs.get("사양", ""),
                            "dimensions": specs.get("치수", ""),
                            "packaging": specs.get("포장", ""),
                            "mold": specs.get("금형", ""),
                            "note": specs.get("기타사항", ""),
                            "images_count": len(data.get("images", [])),
                            "has_print_area": data.get("print_area_url") is not None,
                            "print_area_url": data.get("print_area_url", ""),
                            "cost": "",
                            "price": "",
                            "production": "",
                            "category": self._extract_category_from_path(json_file)
                        }
                except Exception as e:
                    logger.warning(f"Error loading {json_file}: {e}")

        # Load Excel data
        parser = ExcelParserService()
        excel_products = {}

        for excel_file in parser.raw_dir.glob("*.xlsx"):
            try:
                products_list = parser.parse_excel(excel_file.name)

                for product in products_list:
                    if product.code:
                        excel_products[product.code] = {
                            "source": "excel",
                            "product_name": f"{product.code} {product.spec}",
                            "product_code": product.code,
                            "material": self._extract_material(product.spec, excel_file.name),
                            "spec": product.spec,
                            "dimensions": "",
                            "packaging": product.packaging,
                            "mold": product.mold,
                            "note": product.note,
                            "images_count": len(product.images),
                            "has_print_area": False,
                            "print_area_url": "",
                            "cost": str(product.cost) if product.cost else "",
                            "price": str(product.price) if product.price else "",
                            "production": product.production or "",
                            "category": f"Excel/{self._extract_material(product.spec, excel_file.name)}"
                        }
            except Exception as e:
                logger.error(f"Error loading {excel_file}: {e}")

        logger.info(f"✅ Loaded: {len(crawled_products)} crawled, {len(excel_products)} Excel products")

        return crawled_products, excel_products

    def _extract_category_from_path(self, path: Path) -> str:
        """Extract category from file path"""
        parts = path.parts
        if "Bottle" in parts:
            idx = parts.index("Bottle")
            return f"Bottle/{parts[idx+1]}" if idx+1 < len(parts) else "Bottle"
        elif "Jar" in parts:
            idx = parts.index("Jar")
            return f"Jar/{parts[idx+1]}" if idx+1 < len(parts) else "Jar"
        elif "CapPump" in parts:
            idx = parts.index("CapPump")
            return f"CapPump/{parts[idx+1]}" if idx+1 < len(parts) else "CapPump"
        return "Unknown"

    def _extract_material(self, spec: str, filename: str) -> str:
        """Extract material from spec or filename"""
        spec_upper = spec.upper()
        filename_upper = filename.upper()

        if 'PETG' in spec_upper or 'PETG' in filename_upper:
            return 'PETG'
        elif 'PET' in spec_upper or 'PET' in filename_upper:
            return 'PET'
        elif 'PE' in spec_upper and 'PET' not in spec_upper:
            return 'PE'
        elif 'PP' in spec_upper or 'PP' in filename_upper:
            return 'PP'
        return 'Other'

    def generate_master_comparison_sheet(self, wb, crawled_products, excel_products):
        """Generate master comparison sheet with visual indicators"""

        logger.info("📝 Generating master comparison sheet...")

        ws = wb.create_sheet("Master Comparison", 0)

        # Headers
        headers = [
            "Product Code",
            "Data Source",
            "Product Name (Crawled)",
            "Product Name (Excel)",
            "Material (Crawled)",
            "Material (Excel)",
            "Spec (Crawled)",
            "Spec (Excel)",
            "Spec Match",
            "Dimensions",
            "Category",
            "Packaging",
            "Mold",
            "Cost",
            "Price",
            "Production",
            "Note",
            "Images Count (Crawled)",
            "Images Count (Excel)",
            "Has Print Area",
            "Print Area URL",
            "Data Quality Status"
        ]

        # Write headers with formatting
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = self.LIGHT_GRAY_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.CENTER_ALIGN
            cell.border = self.THIN_BORDER

        # Get all unique product codes
        all_codes = sorted(set(list(crawled_products.keys()) + list(excel_products.keys())))

        row = 2
        for code in all_codes:
            crawled = crawled_products.get(code, {})
            excel = excel_products.get(code, {})

            # Determine data source
            if crawled and excel:
                source = "BOTH"
                source_fill = self.GREEN_FILL
            elif crawled:
                source = "Crawled Only"
                source_fill = self.YELLOW_FILL
            else:
                source = "Excel Only"
                source_fill = self.ORANGE_FILL

            # Determine spec match
            if crawled and excel:
                spec_match = "✓" if crawled.get("spec") == excel.get("spec") else "✗ MISMATCH"
                spec_match_fill = self.GREEN_FILL if spec_match == "✓" else self.RED_FILL
                spec_match_font = self.WHITE_FONT if spec_match != "✓" else self.BOLD_FONT
            else:
                spec_match = "N/A"
                spec_match_fill = None
                spec_match_font = None

            # Data quality status
            issues = []
            if not crawled:
                issues.append("NOT_ON_WEBSITE")
            if not excel:
                issues.append("NOT_IN_EXCEL")
            if crawled and not crawled.get("has_print_area"):
                issues.append("NO_PRINT_AREA")
            if crawled and crawled.get("images_count", 0) == 0:
                issues.append("NO_IMAGES")
            if excel and not excel.get("cost"):
                issues.append("NO_COST")
            if excel and not excel.get("price"):
                issues.append("NO_PRICE")

            quality_status = " | ".join(issues) if issues else "✓ COMPLETE"
            quality_fill = self.RED_FILL if issues else self.GREEN_FILL
            quality_font = self.WHITE_FONT if issues else self.BOLD_FONT

            # Write data row
            data_row = [
                code,
                source,
                crawled.get("product_name", "MISSING"),
                excel.get("product_name", "MISSING"),
                crawled.get("material", "MISSING"),
                excel.get("material", "MISSING"),
                crawled.get("spec", "MISSING"),
                excel.get("spec", "MISSING"),
                spec_match,
                crawled.get("dimensions", "MISSING") if crawled else "MISSING",
                crawled.get("category", "MISSING") if crawled else excel.get("category", "MISSING"),
                crawled.get("packaging", excel.get("packaging", "MISSING")),
                crawled.get("mold", excel.get("mold", "MISSING")),
                excel.get("cost", "MISSING") if excel else "MISSING",
                excel.get("price", "MISSING") if excel else "MISSING",
                excel.get("production", "MISSING") if excel else "MISSING",
                crawled.get("note", excel.get("note", "")),
                crawled.get("images_count", "MISSING") if crawled else "MISSING",
                excel.get("images_count", "MISSING") if excel else "MISSING",
                "✓" if crawled and crawled.get("has_print_area") else "✗",
                crawled.get("print_area_url", "MISSING") if crawled else "MISSING",
                quality_status
            ]

            for col_num, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col_num, value=value)
                cell.alignment = self.LEFT_ALIGN
                cell.border = self.THIN_BORDER

                # Apply color coding
                if col_num == 2:  # Data Source column
                    cell.fill = source_fill
                    if source != "BOTH":
                        cell.font = self.WHITE_FONT
                elif col_num == 9:  # Spec Match column
                    if spec_match_fill:
                        cell.fill = spec_match_fill
                        if spec_match_font:
                            cell.font = spec_match_font
                elif col_num == 22:  # Quality Status column
                    cell.fill = quality_fill
                    cell.font = quality_font
                elif value == "MISSING":
                    cell.fill = self.RED_FILL
                    cell.font = self.WHITE_FONT
                elif value == "✗":
                    cell.fill = self.RED_FILL
                    cell.font = self.WHITE_FONT

            row += 1

        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Freeze header row
        ws.freeze_panes = "A2"

        logger.info(f"✅ Master comparison: {row-2} products")

    def generate_missing_data_sheet(self, wb, crawled_products, excel_products):
        """Generate sheet showing only products with missing data"""

        logger.info("📝 Generating missing data sheet...")

        ws = wb.create_sheet("Missing Data Analysis")

        headers = [
            "Product Code",
            "Issue Type",
            "Missing Field",
            "Available In",
            "Priority",
            "Action Required"
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = self.LIGHT_GRAY_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.CENTER_ALIGN

        row = 2
        all_codes = sorted(set(list(crawled_products.keys()) + list(excel_products.keys())))

        for code in all_codes:
            crawled = crawled_products.get(code, {})
            excel = excel_products.get(code, {})

            # Check for missing data
            if not crawled:
                ws.cell(row=row, column=1, value=code)
                ws.cell(row=row, column=2, value="Product Not on Website").fill = self.RED_FILL
                ws.cell(row=row, column=3, value="All crawled fields")
                ws.cell(row=row, column=4, value="Excel Only")
                ws.cell(row=row, column=5, value="HIGH").fill = self.RED_FILL
                ws.cell(row=row, column=6, value="Check if discontinued or search website")
                row += 1

            if not excel:
                ws.cell(row=row, column=1, value=code)
                ws.cell(row=row, column=2, value="Product Not in Excel").fill = self.ORANGE_FILL
                ws.cell(row=row, column=3, value="Cost, Price, Production")
                ws.cell(row=row, column=4, value="Website Only")
                ws.cell(row=row, column=5, value="MEDIUM").fill = self.YELLOW_FILL
                ws.cell(row=row, column=6, value="Add to Excel master catalog")
                row += 1

            if crawled and not crawled.get("has_print_area"):
                ws.cell(row=row, column=1, value=code)
                ws.cell(row=row, column=2, value="No Print Area").fill = self.YELLOW_FILL
                ws.cell(row=row, column=3, value="Print Area PDF/Image")
                ws.cell(row=row, column=4, value="Crawled Data")
                ws.cell(row=row, column=5, value="LOW").fill = self.GREEN_FILL
                ws.cell(row=row, column=6, value="Check website for print area info")
                row += 1

            if excel and not excel.get("cost"):
                ws.cell(row=row, column=1, value=code)
                ws.cell(row=row, column=2, value="No Cost Data").fill = self.RED_FILL
                ws.cell(row=row, column=3, value="Cost")
                ws.cell(row=row, column=4, value="Excel Data")
                ws.cell(row=row, column=5, value="HIGH").fill = self.RED_FILL
                ws.cell(row=row, column=6, value="Add cost information to Excel")
                row += 1

        # Auto-size
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        ws.freeze_panes = "A2"
        logger.info(f"✅ Missing data analysis: {row-2} issues identified")

    def generate_summary_sheet(self, wb, crawled_products, excel_products):
        """Generate summary dashboard sheet"""

        logger.info("📝 Generating summary dashboard...")

        ws = wb.create_sheet("Quality Summary", 0)

        # Title
        ws.cell(row=1, column=1, value="Data Quality Dashboard").font = Font(size=16, bold=True)

        # Calculate statistics
        all_codes = set(list(crawled_products.keys()) + list(excel_products.keys()))
        common_codes = set(crawled_products.keys()) & set(excel_products.keys())
        crawled_only = set(crawled_products.keys()) - set(excel_products.keys())
        excel_only = set(excel_products.keys()) - set(crawled_products.keys())

        # Spec mismatches
        spec_mismatches = 0
        for code in common_codes:
            if crawled_products[code].get("spec") != excel_products[code].get("spec"):
                spec_mismatches += 1

        # Missing print areas
        no_print_area = sum(1 for p in crawled_products.values() if not p.get("has_print_area"))

        # Missing costs
        no_cost = sum(1 for p in excel_products.values() if not p.get("cost"))

        # Statistics table
        row = 3
        stats = [
            ("Total Unique Products", len(all_codes), None),
            ("Products in Both Sources", len(common_codes), self.GREEN_FILL),
            ("Crawled Only", len(crawled_only), self.YELLOW_FILL),
            ("Excel Only", len(excel_only), self.ORANGE_FILL),
            ("Spec Mismatches", spec_mismatches, self.RED_FILL if spec_mismatches > 0 else None),
            ("Missing Print Areas", no_print_area, self.YELLOW_FILL if no_print_area > 0 else None),
            ("Missing Cost Data", no_cost, self.RED_FILL if no_cost > 0 else None),
        ]

        for stat_name, stat_value, fill in stats:
            ws.cell(row=row, column=1, value=stat_name).font = self.BOLD_FONT
            cell = ws.cell(row=row, column=2, value=stat_value)
            cell.font = Font(size=12, bold=True)
            if fill:
                cell.fill = fill
            row += 1

        # Recommendations
        row += 2
        ws.cell(row=row, column=1, value="Priority Actions:").font = Font(size=14, bold=True)
        row += 1

        actions = [
            f"1. Review {spec_mismatches} spec mismatches in 'Master Comparison' sheet",
            f"2. Investigate {len(excel_only)} products that exist in Excel but not on website",
            f"3. Add {len(crawled_only)} website products to Excel catalog",
            f"4. Check {no_print_area} products for print area information",
            f"5. Update cost data for {no_cost} products"
        ]

        for action in actions:
            ws.cell(row=row, column=1, value=action)
            row += 1

        # Auto-size
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15

        logger.info("✅ Summary dashboard complete")

    def generate(self):
        """Generate complete master comparison workbook"""

        logger.info("🚀 Starting master comparison generation...\n")

        # Load all data
        crawled_products, excel_products = self.load_all_data()

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Generate sheets
        self.generate_summary_sheet(wb, crawled_products, excel_products)
        self.generate_master_comparison_sheet(wb, crawled_products, excel_products)
        self.generate_missing_data_sheet(wb, crawled_products, excel_products)

        # Save workbook
        output_file = self.excel_base / "MASTER_COMPARISON_REVIEW.xlsx"
        wb.save(output_file)

        logger.info(f"\n✅ Master comparison generated: {output_file}")
        logger.info("\n📊 Review Instructions:")
        logger.info("   - RED cells = Missing data that requires action")
        logger.info("   - YELLOW cells = Potential issues to review")
        logger.info("   - GREEN cells = Data is complete and matched")
        logger.info("   - ORANGE cells = Excel-only products (not on website)")

        return str(output_file)


def main():
    """Main execution"""
    generator = MasterComparisonGenerator()
    output_file = generator.generate()

    print(f"\n🎉 Complete! Open this file for review:\n   {output_file}")


if __name__ == "__main__":
    main()
