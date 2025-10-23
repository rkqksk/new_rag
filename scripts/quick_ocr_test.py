#!/usr/bin/env python3
"""
빠른 PaddleOCR 테스트 - 출력 확인용
"""

import sys
import json
from pathlib import Path
from paddleocr import PaddleOCR
from PIL import Image, ImageDraw, ImageFont
import openpyxl
import numpy as np

# UTF-8 강제 설정
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

print("=" * 80)
print("PaddleOCR 빠른 테스트 - 출력 확인")
print("=" * 80)

# 1. 작은 Excel 영역을 이미지로 변환
print("\n[1/4] Excel → 이미지 변환 중...")
excel_file = Path("data/excel_uploads/raw/제품 리스트_1.PE.xlsx")
wb = openpyxl.load_workbook(excel_file, data_only=True)
sheet = wb.active

# 10행만 테스트 (제품 1-2개)
start_row, end_row = 5, 15
start_col, end_col = 5, 12  # E~L 열 (제품 2개)

# 이미지 생성
col_widths = {}
for col in range(start_col, end_col + 1):
    col_letter = openpyxl.utils.get_column_letter(col)
    width = sheet.column_dimensions[col_letter].width or 10
    col_widths[col] = max(width, 8) * 10

row_heights = {}
for row in range(start_row, end_row + 1):
    height = sheet.row_dimensions[row].height or 15
    row_heights[row] = max(height, 15) * 1.5

img_width = int(sum(col_widths.values()))
img_height = int(sum(row_heights.values()))

img = Image.new('RGB', (img_width, img_height), 'white')
draw = ImageDraw.Draw(img)

try:
    font = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial.ttf", 14)
except:
    font = ImageFont.load_default()

# 셀 그리기
y_offset = 0
for row in range(start_row, end_row + 1):
    x_offset = 0
    row_height = row_heights.get(row, 20)
    
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=col)
        col_width = col_widths.get(col, 80)
        
        # 셀 테두리
        draw.rectangle(
            [x_offset, y_offset, x_offset + col_width, y_offset + row_height],
            outline='gray',
            width=1
        )
        
        # 셀 값
        if cell.value:
            text = str(cell.value)[:50]
            draw.text((x_offset + 5, y_offset + 5), text, fill='black', font=font)
        
        x_offset += col_width
    
    y_offset += row_height

# 이미지 저장
output_dir = Path("data/excel_uploads/ocr_test")
output_dir.mkdir(parents=True, exist_ok=True)
test_image_path = output_dir / "test_sample.png"
img.save(test_image_path)
print(f"✅ 이미지 저장: {test_image_path}")
print(f"   크기: {img.width}x{img.height} pixels")

wb.close()

# 2. PaddleOCR 초기화
print("\n[2/4] PaddleOCR 초기화 중...")
ocr = PaddleOCR(lang='korean')
print("✅ PaddleOCR 준비 완료")

# 3. OCR 실행
print("\n[3/4] OCR 실행 중...")
img_array = np.array(img)
result = ocr.predict(img_array)

# 4. 결과 분석 및 출력
print("\n[4/4] OCR 결과 분석")
print("=" * 80)

if not result or not result[0]:
    print("❌ OCR 결과 없음")
    sys.exit(1)

# DEBUG: 결과 구조 확인
print(f"\n🔍 DEBUG - OCR 결과 구조:")
print(f"   result 타입: {type(result)}")
print(f"   result 길이: {len(result) if hasattr(result, '__len__') else 'N/A'}")
if result and len(result) > 0:
    ocr_result = result[0]
    print(f"   result[0] 타입: {type(ocr_result)}")
    print(f"   result[0] 길이: {len(ocr_result) if hasattr(ocr_result, '__len__') else 'N/A'}")
    print(f"   result[0] 속성: {dir(ocr_result)[:10]}")  # 처음 10개 속성만

    # OCRResult 객체의 모든 속성 확인
    all_attrs = [attr for attr in dir(ocr_result) if not attr.startswith('_')]
    print(f"   모든 공개 속성: {all_attrs}")

    # OCRResult의 keys 확인
    print(f"   OCRResult keys: {list(ocr_result.keys())[:5]}")

    # print() 메서드로 구조 출력
    print(f"\n   OCRResult.print() 출력:")
    ocr_result.print()
print()

# 텍스트 추출 결과 - OCRResult에서 직접 접근
texts_found = []
ocr_result = result[0]

# OCRResult에서 필드 직접 추출
rec_texts = ocr_result.get('rec_texts', [])
rec_scores = ocr_result.get('rec_scores', [])
dt_polys = ocr_result.get('dt_polys', [])

print(f"OCR 감지 결과:")
print(f"  텍스트 개수: {len(rec_texts)}")
print(f"  신뢰도 개수: {len(rec_scores)}")
print(f"  다각형 개수: {len(dt_polys)}")
print()

# 각 텍스트에 대해 처리
for idx in range(len(rec_texts)):
    text = rec_texts[idx]
    confidence = rec_scores[idx] if idx < len(rec_scores) else 1.0
    bbox = dt_polys[idx] if idx < len(dt_polys) else []

    # 빈 텍스트 스킵
    if not text or text.strip() == '':
        continue

    # UTF-8 인코딩 확인
    text_utf8 = str(text).encode('utf-8').decode('utf-8')

    # 위치 계산 (bbox는 [[x1,y1], [x2,y2], [x3,y3], [x4,y4]] 형식)
    if len(bbox) >= 4:
        try:
            x_center = sum([float(p[0]) for p in bbox]) / len(bbox)
            y_center = sum([float(p[1]) for p in bbox]) / len(bbox)
        except:
            x_center, y_center = 0, 0
    else:
        x_center, y_center = 0, 0

    texts_found.append({
        'index': idx + 1,
        'text': text_utf8,
        'confidence': float(confidence),
        'position': {'x': int(x_center), 'y': int(y_center)},
        'bbox': bbox.tolist() if hasattr(bbox, 'tolist') else list(bbox)
    })

    print(f"{idx+1:3d}. [{float(confidence):5.2f}] '{text_utf8}'")
    print(f"     위치: x={int(x_center)}, y={int(y_center)}")

print(f"\n총 {len(texts_found)}개 텍스트 감지")

# 결과 JSON 저장
result_file = output_dir / "ocr_result.json"
with open(result_file, 'w', encoding='utf-8') as f:
    json.dump(texts_found, f, ensure_ascii=False, indent=2)
print(f"\n💾 결과 저장: {result_file}")

# 시각화 이미지 생성 (OCR 결과 표시)
print("\n[시각화] OCR 감지 영역 표시 중...")
img_viz = img.copy()
draw_viz = ImageDraw.Draw(img_viz)

for item in texts_found:
    bbox = item['bbox']
    # 빨간 박스로 감지 영역 표시
    points = [(int(p[0]), int(p[1])) for p in bbox]
    draw_viz.polygon(points, outline='red', width=2)
    
    # 텍스트 번호 표시
    draw_viz.text((points[0][0], points[0][1] - 15), 
                  f"#{item['index']}", 
                  fill='red', 
                  font=font)

viz_image_path = output_dir / "test_sample_detected.png"
img_viz.save(viz_image_path)
print(f"✅ 시각화 저장: {viz_image_path}")

print("\n" + "=" * 80)
print("테스트 완료!")
print("=" * 80)
print(f"\n확인할 파일:")
print(f"  1. 원본 이미지: {test_image_path}")
print(f"  2. OCR 감지 영역: {viz_image_path}")
print(f"  3. JSON 결과: {result_file}")
print("\n💡 파일을 열어서 OCR이 제대로 인식했는지 확인하세요!")

