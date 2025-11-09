"""
크롤링 기능 테스트 데모

실제로 작동하는지 확인하는 스크립트
"""

import asyncio
import sys
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent))


async def test_static_crawler():
    """정적 크롤러 테스트 - httpbin.org 사용"""
    print("\n" + "="*60)
    print("🧪 테스트 1: Static Crawler (BeautifulSoup)")
    print("="*60)

    from app.services.crawling.static_crawler import StaticCrawler

    crawler = StaticCrawler()

    # 테스트 1: HTML 페이지
    print("\n📄 HTML 페이지 크롤링 중...")
    result = await crawler.crawl('http://httpbin.org/html')

    print(f"✅ 성공!")
    print(f"   - Status: {result['status_code']}")
    print(f"   - URL: {result['url']}")
    print(f"   - Content length: {len(result['text'])} chars")

    # 테스트 2: JSON API
    print("\n📊 JSON API 크롤링 중...")
    result = await crawler.crawl('http://httpbin.org/json', parse_html=False)

    print(f"✅ 성공!")
    print(f"   - Status: {result['status_code']}")
    print(f"   - Content type: {result['headers'].get('content-type')}")

    # 테스트 3: 링크 추출
    print("\n🔗 링크 추출 중...")
    links = await crawler.extract_links('http://httpbin.org/links/5')

    print(f"✅ 성공! 링크 {len(links)}개 추출:")
    for i, link in enumerate(links[:3], 1):
        print(f"   {i}. {link}")


async def test_robots_handler():
    """robots.txt 핸들러 테스트"""
    print("\n" + "="*60)
    print("🤖 테스트 2: Robots.txt Handler")
    print("="*60)

    from app.services.crawling.robots_handler import (
        RobotsHandler,
        RobotsPolicy,
        check_robots,
        bypass_robots
    )

    # 테스트 사이트: GitHub (robots.txt가 있음)
    test_url = 'https://github.com/explore'

    # 모드 1: 준수
    print("\n📋 모드 1: RESPECT (준수)")
    handler = RobotsHandler(policy=RobotsPolicy.RESPECT)
    can_fetch = await handler.can_fetch(test_url)
    print(f"   - {test_url}")
    print(f"   - 크롤링 가능? {can_fetch}")

    # 모드 2: 무시
    print("\n📋 모드 2: IGNORE (무시)")
    handler = RobotsHandler(policy=RobotsPolicy.IGNORE)
    can_fetch = await handler.can_fetch(test_url)
    print(f"   - {test_url}")
    print(f"   - 크롤링 가능? {can_fetch} (항상 True)")

    # 모드 3: 우회
    print("\n📋 모드 3: BYPASS (Googlebot 위장)")
    handler = RobotsHandler(policy=RobotsPolicy.BYPASS)
    can_fetch = await handler.can_fetch(test_url)
    print(f"   - {test_url}")
    print(f"   - 크롤링 가능? {can_fetch}")

    # 간편 함수
    print("\n📋 간편 함수 테스트")
    allowed = await check_robots(test_url, respect=False)
    print(f"   - check_robots(respect=False): {allowed}")

    allowed = await bypass_robots(test_url)
    print(f"   - bypass_robots(): {allowed}")


async def test_evasion_manager():
    """안티 감지 매니저 테스트"""
    print("\n" + "="*60)
    print("🕵️ 테스트 3: Anti-Detection Manager")
    print("="*60)

    from app.services.crawling.evasion import (
        AntiDetectionManager,
        EvasionConfig,
        RateLimiter
    )

    # 이베이션 매니저 생성
    config = EvasionConfig(
        rotate_user_agent=True,
        randomize_headers=True,
        spoof_referrer=True,
        min_delay=0.5,
        max_delay=1.0
    )

    manager = AntiDetectionManager(config)

    # 헤더 생성
    print("\n📝 랜덤 헤더 생성 (3번):")
    for i in range(3):
        headers = manager.get_headers(url='https://example.com')
        print(f"\n   시도 {i+1}:")
        print(f"   - User-Agent: {headers['User-Agent'][:50]}...")
        print(f"   - Referer: {headers.get('Referer', 'None')}")

    # Rate limiter 테스트
    print("\n⏱️ Rate Limiter 테스트 (3건/5초)")
    limiter = RateLimiter(max_requests=3, time_window=5.0)

    for i in range(5):
        await limiter.acquire()
        current_rate = limiter.get_current_rate()
        print(f"   요청 {i+1}: {current_rate:.2f} req/s")

    # 통계
    print("\n📊 이베이션 매니저 통계:")
    stats = manager.get_stats()
    print(f"   - 총 요청: {stats['request_count']}")
    print(f"   - User-Agent 수: {stats['user_agents_count']}")
    print(f"   - 활성 전략: {', '.join(stats['active_strategies'])}")


async def test_excel_processor():
    """Excel 프로세서 테스트"""
    print("\n" + "="*60)
    print("📊 테스트 4: Advanced Excel Processor")
    print("="*60)

    import tempfile
    import openpyxl
    from app.services.data_processing.excel_processor import (
        AdvancedExcelProcessor,
        process_excel_file
    )

    # 테스트 Excel 파일 생성
    print("\n📝 테스트 Excel 파일 생성 중...")

    with tempfile.TemporaryDirectory() as tmpdir:
        test_file = Path(tmpdir) / 'test_complex.xlsx'

        # 복잡한 Excel 파일 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '제품목록'

        # 헤더 (병합)
        ws.merge_cells('A1:D1')
        ws['A1'] = '제품 카탈로그'
        ws['A1'].font = openpyxl.styles.Font(size=16, bold=True)

        # 컬럼 헤더
        ws['A2'] = '제품명'
        ws['B2'] = '가격'
        ws['C2'] = '수량'
        ws['D2'] = '합계'

        # 데이터
        ws['A3'] = 'PET 용기 500ml'
        ws['B3'] = 1000
        ws['C3'] = 50
        ws['D3'] = '=B3*C3'  # 수식

        ws['A4'] = 'PP 뚜껑'
        ws['B4'] = 200
        ws['C4'] = 100
        ws['D4'] = '=B4*C4'  # 수식

        # 합계 (병합)
        ws.merge_cells('A5:C5')
        ws['A5'] = '총합계'
        ws['D5'] = '=SUM(D3:D4)'  # 수식

        # 스타일
        ws['A1'].fill = openpyxl.styles.PatternFill(
            start_color='FFFF00',
            fill_type='solid'
        )

        wb.save(test_file)
        print(f"✅ 파일 생성: {test_file}")

        # Excel 파일 처리
        print("\n🔍 Excel 파일 처리 중...")
        processor = AdvancedExcelProcessor(str(test_file))
        data = processor.extract_all()

        print(f"✅ 처리 완료!")

        # 메타데이터
        print(f"\n📋 메타데이터:")
        print(f"   - 파일 크기: {data.metadata['file_size']} bytes")
        print(f"   - 시트 수: {data.metadata['sheet_count']}")
        print(f"   - 시트 이름: {', '.join(data.metadata['sheet_names'])}")

        # 시트 정보
        sheet = data.sheets['제품목록']
        print(f"\n📊 시트 '제품목록':")
        print(f"   - 행/열: {sheet.max_row} x {sheet.max_col}")
        print(f"   - 병합 셀: {len(sheet.merged_cells)}개")
        print(f"   - 수식: {len(sheet.formulas)}개")

        # 병합 셀 정보
        print(f"\n🔗 병합 셀:")
        for merged in sheet.merged_cells:
            print(f"   - {merged.range}: '{merged.value}' ({merged.cols_span}열 x {merged.rows_span}행)")

        # 수식 정보
        print(f"\n🧮 수식:")
        for cell, formula in sheet.formulas.items():
            print(f"   - {cell}: {formula}")

        # 데이터 샘플
        print(f"\n📝 데이터 샘플 (처음 3행):")
        for i, row in enumerate(sheet.data[:3], 1):
            print(f"   행{i}: {row}")


async def test_multi_strategy_crawler():
    """멀티 전략 크롤러 테스트"""
    print("\n" + "="*60)
    print("🎯 테스트 5: Multi-Strategy Crawler")
    print("="*60)

    from app.services.crawling.multi_strategy_crawler import (
        MultiStrategyCrawler,
        CrawlMethod,
        CrawlConfig
    )
    from app.services.crawling.evasion import EvasionConfig

    # 크롤러 설정
    config = CrawlConfig(
        method=CrawlMethod.AUTO,
        use_evasion=True,
        evasion_config=EvasionConfig(
            rotate_user_agent=True,
            min_delay=0.5,
            max_delay=1.0
        )
    )

    async with MultiStrategyCrawler(config) as crawler:
        # 자동 감지
        print("\n🔍 크롤링 방법 자동 감지...")
        method = await crawler.detect_method('http://httpbin.org/html')
        print(f"   - 감지된 방법: {method.value}")

        # 크롤링
        print("\n📄 크롤링 중...")
        result = await crawler.crawl('http://httpbin.org/html')

        print(f"✅ 성공!")
        print(f"   - URL: {result['url']}")
        print(f"   - 방법: {result['crawl_method']}")
        print(f"   - 콘텐츠 길이: {len(result['content'])} chars")

        # 통계
        print("\n📊 크롤러 통계:")
        stats = crawler.get_stats()
        print(f"   - 총 요청: {stats['total_requests']}")
        print(f"   - Static: {stats['static_requests']}")
        print(f"   - Dynamic: {stats['dynamic_requests']}")
        print(f"   - 실패: {stats['failed_requests']}")


async def main():
    """모든 테스트 실행"""
    print("\n" + "="*60)
    print("🚀 RAG Enterprise - 크롤링 기능 테스트 데모")
    print("="*60)
    print("\n실제 작동하는지 확인합니다...\n")

    try:
        # 테스트 실행
        await test_static_crawler()
        await test_robots_handler()
        await test_evasion_manager()
        await test_excel_processor()
        await test_multi_strategy_crawler()

        # 성공
        print("\n" + "="*60)
        print("✅ 모든 테스트 성공!")
        print("="*60)
        print("\n크롤링 시스템이 정상 작동합니다! 🎉\n")

    except Exception as e:
        print(f"\n❌ 테스트 실패: {e}")
        import traceback
        traceback.print_exc()
        return 1

    return 0


if __name__ == '__main__':
    exit_code = asyncio.run(main())
    sys.exit(exit_code)
